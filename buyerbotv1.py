
# Requested final message fixes
NOGOD_CUSTOM_EMOJI_ID = "6235336389647407554"
SUPPORT_NOTE_TEXT = "কোন সমস্যার কথা জানাতে দ্বিধাবোধ করবেন না.!\nআমাদের সাথে থাকার জন্য ধন্যবাদ।"


# --- PREMIUM EMOJI UPDATE (preserve user-selected IDs) ---
PREMIUM_EMOJI_IDS = {
    "account": "5836782704686798781",
    "report": "5203993413346680064",
    "sell_account": "5357315181649076022",
    "withdrawal": "6046091733924580871",
    "price_list": "5384520307660840171",
    "support": "6113767082636087326",
    "bkash": "6237975191784266396",
    "nogod": "6235336389647407554",
    "facebook": "5933728045067145810",
    "instagram": "4990082283701535678",
    "withdraw_success": "6190336264940559752",
    "withdraw_amount": "6066456029300788292",
    "withdraw_method": "6237975191784266396",
    "withdraw_request": "5291913773307667021",
    "withdraw_completed": "5951982867255924070",
    "withdraw_thanks": "5474525960143385880",
}
# Withdrawal-success copy:
# WITHDRAWAL SUCCESSFUL
# Amount: {amount} BDT
# Method: {method}
# Request ID: #{request_id}
# Withdrawal Completed
# Your payment has been successfully sent.
# Thank you for using our service.
# Labels/details Amount, Method, Request ID must be bold.
# Support note:
# কোন সমস্যার কথা জানাতে দ্বিধাবোধ করবেন না.!
# আমাদের সাথে থাকার জন্য ধন্যবাদ।
# --- END PREMIUM EMOJI UPDATE ---

import os
import re,io,sqlite3,asyncio,time,hashlib,tempfile,shutil,json
from datetime import datetime,timezone
from pathlib import Path
from aiogram import Bot,Dispatcher,F
from aiogram import BaseMiddleware
from aiogram.filters import Command,CommandStart
from aiogram.types import Message,MessageEntity,ReplyKeyboardMarkup,KeyboardButton as _KeyboardButton,ReplyKeyboardRemove,BufferedInputFile,InlineKeyboardMarkup,InlineKeyboardButton as _InlineKeyboardButton
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from openpyxl import load_workbook,Workbook


# --- Premium colored keyboard button engine (merged from premium reference) ---
BUTTON_EMOJI_IDS = {
    "account": "5836782704686798781",
    "report": "5203993413346680064",
    "sell_account": "5357315181649076022",
    "withdrawal": "6046091733924580871",
    "price_list": "5384520307660840171",
    "support": "6113767082636087326",
    "bkash": "6237975191784266396",
    "nogod": "6235336389647407554",
    "admin": "5350396951407895212",
    "link": "5271604874419647061",
    "status": "5231200819986047254",
    "home": "5416041192905265756",
    "gift_box": "5970074171449808121",
    "delete": "5422557736330106570",
    "refer_btn": "5420396762189831222",
    "get_number_btn": "5382357040008021292",
    "cross": "5420130255174145507",
    "stop": "5956074558044770726",
    "ban": "5420323339723881652",
    "done": "6298670698948724690",
    "rocket": "5346042941196507141",
    "binance": "5348212415077064131",
    "live": "5355102594886833928",
    "channel": "6215074610845585917",
    "waiting": "6217721388736712699",
    "back": "5267490665117275176",
    "leader_board": "5280769763398671636",
    "money": "6233367447789899509",
    "change_number": "5402186569006210455",
    "msg": "5253742260054409879",
    "setting": "5341715473882955310",
    "add": "5397916757333654639",
    # Admin/control-panel Premium Emoji IDs supplied by the user.
    "admin_user_management": "5341715473882955310",
    "admin_work_management": "5341715473882955310",
    "admin_force_join": "6206497372176913599",
    "admin_pending_file": "5262838597060422237",
    "admin_generate_excel": "5298853345241358103",
    "admin_payment_channel": "5377336227533969892",
    "admin_receiver_channel": "5350505622670435563",
    "admin_broadcast": "4992560350982309130",
    "admin_new_button": "5195113117490621021",
    "admin_database": "5470016867252855402",
    "admin_maintenance": "5362074752737358235",
    "admin_support": "6206495649895028694",
    "admin_user_list": "5330363643391393348",
    "admin_blocked_user": "5217877998937595307",
    "admin_remove_balance": "5850580158276177819",
    "admin_back": "5341715473882955310",
    "admin_add_work": "5258334778389710253",
    "admin_disable_notice": "5436322999139574951",
    "admin_work_list": "5201940333079846676",
    "admin_add_withdrawal_category": "5244495762401810280",
    "admin_pending_withdrawal": "5262838597060422237",
    "admin_add_force_join_channel": "5244495762401810280",
    "admin_force_join_list": "5201940333079846676",
    "admin_remove_force_join": "6206108815075579644",
    "admin_add_new_button": "5244495762401810280",
    "admin_view_added_button": "5201940333079846676",
    "admin_edit_added_button": "5341715473882955310",
    "admin_download_database": "5435977112538324435",
    "admin_upload_database": "5258134813302332906",
    "admin_delete_all_data": "6206108815075579644",
    "admin_maintenance_on": "5215313353706057331",
    "admin_maintenance_off": "6113685078825505075",
    "admin_withdrawal_management": "5341715473882955310",
    "admin_user_menu": "5341715473882955310",
    "admin_delete_category": "6206108815075579644",
    "admin_add_support": "5244495762401810280",
    "admin_delete_support": "6206108815075579644",
    "admin_support_list": "5201940333079846676",
    "admin_delete_button": "6206108815075579644",
    "admin_upload_report": "5298853345241358103",
}

def _button_meta(text):
    t = re.sub(r"<[^>]+>", "", str(text)).strip().lower()
    rules = [
        # Admin/control-panel buttons: exact labels use the Premium Emoji IDs
        # supplied for this version. Keep these rules before generic rules so
        # labels such as "Maintenance ON" do not inherit another icon.
        (("user management",), "admin_user_management", "primary"),
        (("work management",), "admin_work_management", "success"),
        (("force join",), "admin_force_join", "success"),
        (("pending files", "pending file"), "admin_pending_file", "success"),
        (("generate main excel",), "admin_generate_excel", "danger"),
        (("set payment channel",), "admin_payment_channel", "primary"),
        (("set receiver channel", "set reciver channel"), "admin_receiver_channel", "primary"),
        (("broadcast",), "admin_broadcast", "success"),
        (("new button",), "admin_new_button", "success"),
        (("database",), "admin_database", "danger"),
        (("maintenance",), "admin_maintenance", "danger"),
        (("support",), "admin_support", "danger"),
        (("all user list", "new user list", "user list"), "admin_user_list", "primary"),
        (("blocked user", "blickted user"), "admin_blocked_user", "danger"),
        (("remove balance",), "admin_remove_balance", "danger"),
        (("back to admin panel",), "admin_back", "danger"),
        (("add work",), "admin_add_work", "success"),
        (("disable notice",), "admin_disable_notice", "primary"),
        (("work list",), "admin_work_list", "success"),
        (("add withdrawal category",), "admin_add_withdrawal_category", "success"),
        (("pending withdrawal requests",), "admin_pending_withdrawal", "success"),
        (("add force join channel",), "admin_add_force_join_channel", "success"),
        (("force join channel list",), "admin_force_join_list", "success"),
        (("remove force join channel",), "admin_remove_force_join", "danger"),
        (("add new button",), "admin_add_new_button", "success"),
        (("view added button",), "admin_view_added_button", "success"),
        (("edit added button",), "admin_edit_added_button", "success"),
        (("download database",), "admin_download_database", "success"),
        (("upload database",), "admin_upload_database", "success"),
        (("ডিলিট অল ডাটা", "delete all data"), "admin_delete_all_data", "danger"),
        (("maintenance on",), "admin_maintenance_on", "danger"),
        (("maintenance off", "maintains off"), "admin_maintenance_off", "success"),
        (("support management",), "admin_support", "primary"),
        (("withdrawal management",), "admin_withdrawal_management", "success"),
        (("user menu",), "admin_user_menu", "primary"),
        (("pending withdrawals",), "admin_pending_withdrawal", "primary"),
        (("delete category",), "admin_delete_category", "danger"),
        (("add support button",), "admin_add_support", "success"),
        (("delete support button",), "admin_delete_support", "danger"),
        (("support list",), "admin_support_list", "primary"),
        (("delete button",), "admin_delete_button", "danger"),
        (("view added button",), "admin_view_added_button", "success"),
        (("edit added button",), "admin_edit_added_button", "success"),
        (("upload report file",), "admin_upload_report", "danger"),
        (("add withdrawal category",), "admin_add_withdrawal_category", "success"),
        (("add force join channel",), "admin_add_force_join_channel", "success"),
        (("remove force join channel",), "admin_remove_force_join", "danger"),
        (("add new button",), "admin_add_new_button", "success"),
        (("view added button",), "admin_view_added_button", "success"),
        (("edit added button",), "admin_edit_added_button", "success"),
        (("download database",), "admin_download_database", "success"),
        (("upload database",), "admin_upload_database", "success"),
        (("delete all data",), "admin_delete_all_data", "danger"),
        (("maintenance on",), "admin_maintenance_on", "danger"),
        (("maintenance off",), "admin_maintenance_off", "danger"),
        (("support", "contact support", "support team"), "support", "danger"),
        (("report", "upload report", "future reports"), "report", "primary"),
        # Specific payment/account labels MUST come before the generic "account" rule.
        (("sell account",), "sell_account", "success"),
        (("bkash", "বিকাশ"), "bkash", "primary"),
        (("nagad", "nogad", "নগদ"), "nogad", "primary"),
        (("account",), "account", "primary"),
        (("get number",), "get_number_btn", "primary"),
        (("withdraw",), "withdrawal", "success"),
        (("price list",), "price_list", "primary"),
        (("balance", "all user balance", "add balance"), "money", "success"),
        (("refer",), "refer_btn", "success"),
        (("leaderboard",), "leader_board", "primary"),
        (("admin panel", "user management", "force join", "management", "system config", "security", "broadcast", "b-cast", "user menu"), "admin", "primary"),
        (("cancel", "delete", "remove", "reject", "disable", "block", "ban", "back"), "cross", "danger"),
        (("approve", "confirm", "add", "enable", "unban", "success"), "done", "success"),
        (("setting", "edit", "view", "search", "status", "list", "check joined", "maintenance", "মেইনটেন্যান্স"), "setting", "primary"),
        (("link", "channel", "join"), "link", "primary"),
    ]
    for words, key, style in rules:
        if any(w in t for w in words):
            return BUTTON_EMOJI_IDS.get(key), style
    return BUTTON_EMOJI_IDS.get("msg"), "primary"



def _remove_normal_emojis(text):
    if text is None:
        return text
    return re.sub(r"[\U0001F000-\U0001FAFF\u2600-\u27BF\u2300-\u23FF]", "", str(text)).strip()

def kbbtn(text, **kwargs):
    text = _remove_normal_emojis(text)
    icon, style = _button_meta(text)
    kwargs.setdefault("icon_custom_emoji_id", icon)
    kwargs.setdefault("style", style)
    try:
        return _KeyboardButton(text=text, **kwargs)
    except Exception:
        # Older aiogram versions may reject these Bot API fields with a
        # Pydantic ValidationError rather than TypeError. Fall back to the
        # standard Telegram button so navigation never fails silently.
        kwargs.pop("icon_custom_emoji_id", None)
        kwargs.pop("style", None)
        return _KeyboardButton(text=text, **kwargs)

def ibtn(text, **kwargs):
    text = _remove_normal_emojis(text)
    icon, style = _button_meta(text)
    kwargs.setdefault("icon_custom_emoji_id", icon)
    kwargs.setdefault("style", style)
    try:
        return _InlineKeyboardButton(text=text, **kwargs)
    except Exception:
        kwargs.pop("icon_custom_emoji_id", None)
        kwargs.pop("style", None)
        return _InlineKeyboardButton(text=text, **kwargs)
TOKEN = "8955454814:AAF9dJi46wUi5u0hDgY2rg3wGxP_ONhSg4Y"
ADMIN_ID = 6995426618
DB=os.getenv("DB_PATH","bot.db")
FD=Path(os.getenv("FILES_DIR","uploads")); FD.mkdir(parents=True,exist_ok=True)
if not TOKEN: raise RuntimeError("BOT_TOKEN is missing")
if not ADMIN_ID: raise RuntimeError("ADMIN_ID is missing")

db=sqlite3.connect(DB,check_same_thread=False); db.row_factory=sqlite3.Row
c=db.cursor()
c.executescript("""
CREATE TABLE IF NOT EXISTS users(user_id INTEGER PRIMARY KEY,username TEXT DEFAULT '',balance REAL DEFAULT 0,total_income REAL DEFAULT 0,today_income REAL DEFAULT 0,today_date TEXT DEFAULT '',created_at TEXT DEFAULT '',blocked INTEGER DEFAULT 0);
CREATE TABLE IF NOT EXISTS categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,price REAL DEFAULT 0,active INTEGER DEFAULT 1);
CREATE TABLE IF NOT EXISTS submissions(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,username TEXT,category_id INTEGER,category_name TEXT,price REAL,rows_count INTEGER,estimated_profit REAL,file_name TEXT,file_path TEXT,status TEXT DEFAULT 'pending',created_at TEXT);
CREATE TABLE IF NOT EXISTS withdrawal_categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,minimum REAL DEFAULT 0,maximum REAL DEFAULT 0,fee REAL DEFAULT 0,active INTEGER DEFAULT 1);
CREATE TABLE IF NOT EXISTS withdrawals(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,username TEXT,category_id INTEGER,category_name TEXT,amount REAL,wallet TEXT,fee REAL DEFAULT 0,total_debited REAL DEFAULT 0,status TEXT DEFAULT 'pending',created_at TEXT);
CREATE TABLE IF NOT EXISTS support(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT,text TEXT DEFAULT '',updated_at TEXT);
CREATE TABLE IF NOT EXISTS support_buttons(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,url TEXT NOT NULL,active INTEGER DEFAULT 1,created_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS force_join_channels(id INTEGER PRIMARY KEY AUTOINCREMENT,chat_id TEXT UNIQUE,channel_name TEXT,invite_link TEXT,active INTEGER DEFAULT 1,created_at TEXT);
CREATE TABLE IF NOT EXISTS bot_settings(key TEXT PRIMARY KEY,value TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS file_reviews(submission_id INTEGER PRIMARY KEY,accepted_count INTEGER DEFAULT 0,rejected_count INTEGER DEFAULT 0,final_amount REAL DEFAULT 0,note TEXT DEFAULT '',status TEXT DEFAULT 'reviewing',updated_at TEXT DEFAULT '',admin_file_message_id INTEGER DEFAULT 0,pending_list_message_id INTEGER DEFAULT 0,review_prompt_message_id INTEGER DEFAULT 0);
CREATE TABLE IF NOT EXISTS custom_buttons(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE,target TEXT,active INTEGER DEFAULT 1,created_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS reports(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_name TEXT NOT NULL,
    file_path TEXT NOT NULL,
    uploaded_at REAL NOT NULL,
    expires_at REAL NOT NULL,
    green_count INTEGER DEFAULT 0,
    processed_users INTEGER DEFAULT 0,
    total_reward REAL DEFAULT 0,
    category_name TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS paid_uids(
    uid TEXT PRIMARY KEY,
    user_id INTEGER NOT NULL,
    submission_id INTEGER NOT NULL,
    report_id INTEGER NOT NULL,
    amount REAL DEFAULT 0,
    paid_at TEXT NOT NULL
);
"""); db.commit()
# Keep the selected work/category with each uploaded report.
try:
    c.execute("ALTER TABLE reports ADD COLUMN category_name TEXT DEFAULT ''")
    db.commit()
except sqlite3.OperationalError:
    pass
try:
    c.execute("ALTER TABLE users ADD COLUMN blocked INTEGER DEFAULT 0")
    db.commit()
except sqlite3.OperationalError:
    pass
# Keep a human-readable account name for payment approval reports.
for _table, _col, _definition in [
    ("users", "full_name", "TEXT DEFAULT ''"),
    ("withdrawals", "full_name", "TEXT DEFAULT ''"),
]:
    try:
        c.execute(f"ALTER TABLE {_table} ADD COLUMN {_col} {_definition}")
        db.commit()
    except sqlite3.OperationalError:
        pass

# Store a SHA-256 fingerprint for each uploaded XLSX so the same user cannot submit the same file twice.
try:
    c.execute("ALTER TABLE submissions ADD COLUMN file_hash TEXT DEFAULT ''")
    db.commit()
except sqlite3.OperationalError:
    pass

# Persist Telegram's file_id and receiver message id so Pending Files does not
# depend on Railway's ephemeral local filesystem.
for _table, _col, _definition in [
    ("submissions", "file_id", "TEXT DEFAULT ''"),
    ("submissions", "receiver_message_id", "INTEGER DEFAULT 0"),
]:
    try:
        c.execute(f"ALTER TABLE {_table} ADD COLUMN {_col} {_definition}")
        db.commit()
    except sqlite3.OperationalError:
        pass

for _col, _definition in [
    ("admin_file_message_id", "INTEGER DEFAULT 0"),
    ("pending_list_message_id", "INTEGER DEFAULT 0"),
    ("review_prompt_message_id", "INTEGER DEFAULT 0"),
]:
    try:
        c.execute(f"ALTER TABLE file_reviews ADD COLUMN {_col} {_definition}")
        db.commit()
    except sqlite3.OperationalError:
        pass



# --- Telegram Premium Custom Emoji support ---
PE = {
    "welcome": "5354972242629383937", "welcome_action": "5809671801667589890", "welcome_down": "6206172741368813487",
    "account": "5841276284155467413", "user_id": "6269073697059901810", "balance": "5215420556089776398",
    "today_income": "5256186332669035163", "total_income": "5391316230118321870", "report": "5203993413346680064",
    "coming_soon": "5974235702701853774", "your_accounts": "6235699516247379290", "estimated_profit": "6104764195528511604",
    "work_prompt": "5931793690581275367", "xlsx": "5319080489426887527", "work_selected": "5276481311568043264",
    "price": "5854965264050818921", "warning_1": "5244484741515732344", "warning_2": "5247187233722607160",
    "xlsx_now": "5352782831510635024", "unlabeled_1": "4970019434899964604", "receiver_warning": "5366417913861397904",
    "file_success": "5354972242629383937", "submit_work": "5316688029434264397", "submit_account": "5936017305585586269",
    "submit_price": "5404635941610463865", "submit_profit": "6206503415195899956", "admin_verify": "6298747815086524010",
    "unlabeled_2": "5215522595922779944", "price_list": "5195450074854865713", "price_item": "6206245785877616415",
    "support": "6206495649895028694", "support_note": "6113904371265704130", "withdraw_prompt": "5472250091332993630",
    "bkash": "6237975191784266396", "nagad": "6235336389647407554", "minmax": "5350512473143273043",
    "amount_error": "5462998469784393745", "balance_error": "5816566289329556959", "wallet_bkash": "6237975191784266396",
    "withdraw_success": "5355142851615283756", "request_id": "5291913773307667021", "amount": "6066456029300788292",
    "report_category": "5357315181649076022", "report_check": "5235795597474157178",
}
# Correct the withdrawal prompt ID if the saved mapping was accidentally mistyped.
PE["withdraw_prompt"] = "5472250091332993630"

# Only these ordinary Unicode emojis are converted. The ordinary glyph is REMOVED,
# and exactly one custom-emoji placeholder is emitted in its place. This prevents
# the old "normal + premium" double-emoji problem.
UNICODE_TO_PE = {
    # 1-9: intentionally unchanged because no new IDs were supplied.
    "🥰": PE["welcome"], "💎": PE["welcome_action"], "🔽": PE["welcome_down"],
    "📊": PE["report"], "⏳": PE["coming_soon"], "👤": PE["your_accounts"],
    "💰": PE["balance"], "📁": PE["unlabeled_1"], "📞": PE["support"],

    # User-supplied Premium Emoji IDs (serial 10-34).
    "💳": "6305298855688672996",
    "📂": "5240478589295552614",
    "💵": "5224257782013769471",
    "📈": "6154657967317717391",
    "📱": "5246723905535632915",
    "⭐": "6136464120779638846",
    "🔧": "5471893218205392288",
    "✓": "5386448185336017745",
    "🔐": "5291873529464122510",
    "✅": "5780463361175066565",
    "📌": "5280657037687016089",
    "🟢": "6113685078825505075",
    "🔴": "5215313353706057331",
    "➕": "6156854417887859761",
    "❌": "5462998469784393745",
    "✏️": "6204162490515855272",
    "🔕": "5260264520080695245",
    "🔗": "5388583647370565067",
    "⚠️": "5213300332599192534",
    "🗑️": "5382116818192182763",
    "📋": "5821051579511345616",
    "🗑": "5382116818192182763",
    "🗄": "4967762670104085632",
    "📤": "4967897119760319376",
    "📚": "6147871412183899488",

    # 35 (👥): intentionally unchanged because no new ID was supplied.
    "📢": "5399967660052081305",
    "🔍": "6206446249181189526",
    # 38-43 occur only in the internal color-strip regex; their literal glyphs
    # are removed from source below, so they are not used as display emojis.
}

# Text phrases that receive one custom emoji BEFORE the phrase, while keeping the phrase.
PHRASE_PE = [
    # Category-specific Premium/Custom Emoji: every visible occurrence gets the matching emoji.
    ("Facebook", PREMIUM_EMOJI_IDS["facebook"]),
    ("Instagram", PREMIUM_EMOJI_IDS["instagram"]),
    ("আপনার file সফলভাবে জমা হয়েছে.", PE["file_success"]),
    ("Admin যাচাই করার পর Balance-এ টাকা যোগ হবে.", PE["admin_verify"]),
    ("Admin যাচাই করার পর Balance-এ টাকা যোগ হবে.", PE["admin_verify"]),
    ("Withdrawal request জমা হয়েছে.", PE["withdraw_success"]),
    ("Admin approval-এর অপেক্ষায়.", PE["admin_verify"]),
    ("Receiver Channel এখনো সেট করা হয়নি।", PE["receiver_warning"]),
    ("Withdrawal category নির্বাচন করুন:", PE["withdraw_prompt"]),
    ("আপনার বিকাশ নাম্বার দিন।", PE["wallet_bkash"]), ("আপনার Nogod নাম্বার দিন।", PE["nagad"]),
    ("সঠিক amount লিখুন।", PE["amount_error"]),
    ("রিপোর্টে ভুল আসলে", PE["warning_1"]),
    ("Report বাটনে ক্লিক করে আপনার রিপোর্ট চেক করতে পারেন।", PE["report_check"]),
    ("Report Available", PE["report"]), ("রিপোর্ট এসে গেছে!", PE["report"]),
    ("Price List", PE["price_list"]), ("Your Accounts:", PE["your_accounts"]),
    ("Estimated Profit:", PE["estimated_profit"]), ("Coming Soon...", PE["coming_soon"]),
    ("Work selected:", PE["work_selected"]), ("কাজ নির্বাচন করুন।", PE["work_prompt"]),
    ("XLSX file", PE["xlsx"]), ("গ্রাহক সেবা কেন্দ্র", PE["support"]),     ("Category: Bkash", PE["bkash"]), ("Category: Nogod", PE["nagad"]), ("Category:", PE["report_category"]),
    ("User ID:", PE["user_id"]), ("Today's Income:", PE["today_income"]), ("Total Income:", PE["total_income"]),
    ("Balance:", PE["balance"]), ("Request ID:", PE["request_id"]), ("Amount:", PE["amount"]),
    ("Work:", PE["submit_work"]), ("Price:", PE["submit_price"]), ("Account:", PE["submit_account"]),
    ("Estimated Profit:", PE["submit_profit"]),
    ("Minimum:", PE["minmax"]), ("Maximum:", PE["minmax"]),
]

GENERIC_WORD_PE = [
    (r"(?<!\w)Account(?!\w)", PE["account"]),
    (r"(?<!\w)Report(?!\w)", PE["report"]),
    (r"(?<!\w)Balance(?!\w)", PE["balance"]),
    (r"(?<!\w)Price(?!\w)", PE["price"]),
    (r"(?<!\w)Amount(?!\w)", PE["amount"]),
    (r"(?<!\w)Work(?!\w)", PE["submit_work"]),
    (r"(?<!\w)XLSX(?!\w)", PE["xlsx"]),
    (r"(?<!\w)Support(?!\w)", PE["support"]),
    (r"(?<!\w)Withdrawal(?!\w)", PE["withdraw_prompt"]),
]

BOLD_LABELS = [
    "Account:", "User ID:", "Balance:", "Today's Income:", "Total Income:",
    "Your Accounts:", "Estimated Profit:", "Work:", "Accounts:", "Price:",
    "Category:", "Minimum:", "Maximum:", "Request ID:", "Amount:",
    "Processed Users:", "Total Reward:", "Report:", "Report expires in:",
    "Final Amount:", "Added Balance:", "Withdrawal Amount:", "Payment Number:",
]

def _u16_len(t):
    return len(t.encode("utf-16-le")) // 2

def premiumize(text):
    if text is None:
        return text, None
    # Legacy withdrawal prompts sometimes contained a normal phone/star emoji
    # before the Premium Custom Emoji. Strip those glyphs first.
    original = str(text)
    original = original.replace("📱 ⭐ আপনার বিকাশ নাম্বার দিন।", "আপনার বিকাশ নাম্বার দিন।")
    original = original.replace("📱 ⭐ আপনার Nogod নাম্বার দিন।", "আপনার Nogod নাম্বার দিন।")
    original = original.replace("📱 আপনার বিকাশ নাম্বার দিন।", "আপনার বিকাশ নাম্বার দিন।")
    original = original.replace("📱 আপনার Nogod নাম্বার দিন।", "আপনার Nogod নাম্বার দিন।")

    # 1) Remove/replace ordinary emoji FIRST, so a phrase rule cannot create a second emoji.
    # Keep a placeholder marker that is later converted to a Telegram custom-emoji entity.
    s = original
    markers = []
    for ch, eid in sorted(UNICODE_TO_PE.items(), key=lambda x: len(x[0]), reverse=True):
        if ch in s:
            s = s.replace(ch, "\uFFF0")
            # Record all markers with their custom emoji IDs in sequence.
            markers.extend([eid] * original.count(ch))

    # Remove every remaining ordinary Unicode emoji. Custom emojis are represented
    # by Telegram MessageEntity and are not affected by this cleanup.
    _remaining_emoji_re = re.compile(
        r"[\U0001F000-\U0001FAFF\u2600-\u27BF\u2300-\u23FF]"
    )
    s = _remaining_emoji_re.sub("", s)

    # 2) Insert a marker before phrases that don't already have a marker immediately before them.
    # Process longest phrases first.
    phrase_matches = []
    for phrase, eid in sorted(PHRASE_PE, key=lambda x: len(x[0]), reverse=True):
        for m in re.finditer(re.escape(phrase), s):
            before = s[:m.start()]
            # If a marker is already directly before the phrase (ignoring one space), don't add another.
            j = len(before.rstrip())
            if j > 0 and before[j-1] == "\uFFF0":
                continue
            phrase_matches.append((m.start(), eid))
    # Avoid multiple phrase insertions at the same/overlapping point.
    used = set()
    inserts = []
    for pos, eid in sorted(phrase_matches, key=lambda x: (x[0], -len(next((p for p,e in PHRASE_PE if e==x[1]), "")))):
        if pos in used:
            continue
        used.add(pos); inserts.append((pos, eid))
    for pos, eid in reversed(inserts):
        s = s[:pos] + "\uFFF0" + s[pos:]

    # 3) Build text + exact custom-emoji entities. Every marker becomes exactly one entity.
    out=[]; entities=[]; marker_index=0
    for ch in s:
        if ch == "\uFFF0":
            offset = _u16_len("".join(out))
            # Find the correct ID by replaying marker IDs in textual order.
            # If phrase insertions exceed the original marker list, derive it from nearby phrase markers below.
            eid = None
            # We'll resolve IDs from a parallel pass instead.
            out.append("⭐")
        else:
            out.append(ch)
    rendered="".join(out)

    # Resolve marker IDs by replaying replacements/phrase insertions on a second pass.
    # This is deterministic and avoids relying on Python string offsets after insertions.
    id_seq=[]
    # Original emoji markers in order
    tmp=original
    for ch in tmp:
        if ch in UNICODE_TO_PE:
            id_seq.append(UNICODE_TO_PE[ch])
    # The above is not enough when phrase markers are interleaved, so map by scanning
    # the rendered marker positions against the original/phrase construction again.
    # Reconstruct marker IDs exactly.
    work=original
    seq=[]
    for ch,eid in sorted(UNICODE_TO_PE.items(), key=lambda x: len(x[0]), reverse=True):
        # replacement is global; preserve sequence by storing temporary unique tokens instead.
        pass
    # Simpler robust reconstruction: redo transformation using unique tokens.
    token_text=original
    token_ids=[]; token_num=0
    for ch,eid in sorted(UNICODE_TO_PE.items(), key=lambda x: len(x[0]), reverse=True):
        token=f"\uFFF1{token_num}\uFFF2"; token_num+=1
        count=token_text.count(ch)
        if count:
            token_text=token_text.replace(ch, token)
            token_ids.extend([eid]*count)
    # Rebuild token sequence in actual order by scanning token occurrences is hard with one token per emoji type;
    # use a regex over all token patterns.
    token_map={}
    token_num=0
    token_text=original

    # Direct Premium Custom Emoji markers: {{PE:<custom_emoji_id>}}
    # These are used for the withdrawal success UI so the exact user-provided
    # Premium Emoji IDs are rendered without leaving a normal Unicode emoji.
    direct_pe = re.compile(r"\{\{PE:(\d+)\}\}")
    for m in list(direct_pe.finditer(token_text)):
        token=f"\uFFF3{len(token_map)}\uFFF4"
        token_map[token]=m.group(1)
        token_text=token_text.replace(m.group(0), token, 1)
    for ch,eid in sorted(UNICODE_TO_PE.items(), key=lambda x: len(x[0]), reverse=True):
        token=f"\uFFF1{token_num}\uFFF2"; token_num+=1
        token_text=token_text.replace(ch, token)
        token_map[token]=eid
    # phrase inserts on token_text
    for phrase,eid in sorted(PHRASE_PE, key=lambda x: len(x[0]), reverse=True):
        start=0
        while True:
            idx=token_text.find(phrase,start)
            if idx<0: break
            before=token_text[:idx].rstrip()
            if before.endswith(tuple(token_map.keys())):
                start=idx+len(phrase)
                continue
            token=f"\uFFF3{len(token_map)}\uFFF4"
            token_map[token]=eid
            token_text=token_text[:idx]+token+token_text[idx:]
            start=idx+len(token)+len(phrase)

    # standalone English headings without a Unicode emoji (e.g. plain `Account`)
    # also receive one custom emoji. If a mapped token is already immediately before
    # the word, do not add another.
    for pattern,eid in GENERIC_WORD_PE:
        start=0
        while True:
            m=re.search(pattern, token_text[start:])
            if not m: break
            idx=start+m.start(); end=start+m.end()
            before=token_text[:idx].rstrip()
            if before.endswith(tuple(token_map.keys())):
                start=end
                continue
            token=f"\uFFF3{len(token_map)}\uFFF4"
            token_map[token]=eid
            token_text=token_text[:idx]+token+token_text[idx:]
            start=idx+len(token)+len(m.group(0))

    rendered_parts=[]; ids=[]; i=0
    # tokenize custom markers and produce ⭐ placeholders
    marker_pattern=re.compile(r"\uFFF1\d+\uFFF2|\uFFF3\d+\uFFF4")
    for m in marker_pattern.finditer(token_text):
        rendered_parts.append(token_text[i:m.start()])
        if m.group(0).startswith("\uFFF3"):
            rendered_parts.append("⭐ ")
        else:
            # Replaced ordinary emoji: add a space only when the original emoji
            # was directly attached to text (e.g. a custom emoji followed by text).
            nxt = token_text[m.end():m.end()+1]
            rendered_parts.append("⭐ " if nxt and not nxt.isspace() else "⭐")
        ids.append(token_map[m.group(0)]); i=m.end()
    rendered_parts.append(token_text[i:])
    rendered="".join(rendered_parts)

    for idx,eid in enumerate(ids):
        # Find idx-th ⭐ in rendered.
        pos=0
        for _ in range(idx+1):
            pos=rendered.find("⭐", pos)+1
        start=pos-1
        entities.append(MessageEntity(type="custom_emoji", offset=_u16_len(rendered[:start]), length=1, custom_emoji_id=str(eid)))

    # 4) Bold the requested labels without disturbing custom-emoji offsets.
    for label in BOLD_LABELS:
        for m in re.finditer(re.escape(label), rendered):
            entities.append(MessageEntity(type="bold", offset=_u16_len(rendered[:m.start()]), length=_u16_len(m.group(0))))

    return rendered, entities or None

async def ans(target, text, **kwargs):
    text, entities = premiumize(text)
    if entities is not None and "entities" not in kwargs:
        kwargs["entities"] = entities
        kwargs["parse_mode"] = None
    return await target.answer(text, **kwargs)

async def sendmsg(bot, chat_id, text, **kwargs):
    text, entities = premiumize(text)
    if entities is not None and "entities" not in kwargs:
        kwargs["entities"] = entities
        kwargs["parse_mode"] = None
    return await bot.send_message(chat_id, text, **kwargs)

async def senddoc(bot, chat_id, document, **kwargs):
    if "caption" in kwargs and kwargs["caption"] is not None:
        cap, entities = premiumize(kwargs["caption"])
        kwargs["caption"] = cap
        if entities is not None and "caption_entities" not in kwargs:
            kwargs["caption_entities"] = entities
            kwargs["parse_mode"] = None
    return await bot.send_document(chat_id, document, **kwargs)

def cleanup_expired_report():
    """Remove every report whose 18-hour lifetime has ended."""
    rows=c.execute("SELECT id,file_path FROM reports WHERE expires_at<=?", (time.time(),)).fetchall()
    if not rows:
        return
    for r in rows:
        try:
            Path(r["file_path"]).unlink(missing_ok=True)
        except Exception:
            pass
        c.execute("DELETE FROM reports WHERE id=?", (r["id"],))
    db.commit()

def normalize_uid(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value) if value >= 0 else None
    if isinstance(value, float):
        if value.is_integer() and value >= 0:
            return str(int(value))
        return None
    s=str(value).strip()
    if not s or s in {"10/7"}:
        return None
    # Only real numeric UID values are accepted.
    if not re.fullmatch(r"\d+", s):
        return None
    return s

def _cell_looks_green(cell):
    # Detect common Excel green fills/fonts, including RGB/indexed/theme variants.
    for obj_name in ("fill", "font"):
        try:
            obj=getattr(cell, obj_name)
            color = getattr(obj, "fgColor", None) if obj_name == "fill" else getattr(obj, "color", None)
            if not color:
                continue
            rgb=getattr(color, "rgb", None)
            if isinstance(rgb, str):
                rr=rgb[-6:].upper()
                if len(rr)==6:
                    r1,g1,b1=int(rr[0:2],16),int(rr[2:4],16),int(rr[4:6],16)
                    if g1 >= 100 and g1 > r1*1.20 and g1 > b1*1.10:
                        return True
            indexed=getattr(color, "indexed", None)
            if indexed in {4, 9, 10, 11, 35, 43, 51}:
                return True
        except Exception:
            pass
    return False

def row_is_green(row):
    # A report row is valid when a cell explicitly contains Green/Valid
    # (case-insensitive, including phrases such as "Green - Valid"),
    # or a cell has a clearly green Excel fill/font.
    words=("green", "valid", "verified", "active", "success", "successful", "live")
    for cell in row:
        v=cell.value
        if isinstance(v,str):
            t=v.strip().lower()
            if any(w in t for w in words):
                return True
        if _cell_looks_green(cell):
            return True
    return False

def _find_uid_column(ws):
    # Prefer a column whose header contains UID; fall back to column A.
    try:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 10)):
            for cell in row:
                v=cell.value
                if isinstance(v,str) and "uid" in v.strip().lower():
                    return cell.column
    except Exception:
        pass
    return 1

def extract_report_green_uids(raw):
    green=set()
    wb=load_workbook(io.BytesIO(raw),read_only=False,data_only=True)
    for ws in wb.worksheets:
        uid_col=_find_uid_column(ws)
        for row in ws.iter_rows():
            vals=list(row)
            if not vals or uid_col > len(vals):
                continue
            uid=normalize_uid(vals[uid_col-1].value)
            if uid and row_is_green(vals):
                green.add(uid)
    wb.close()
    return green

def extract_submission_uids(file_path):
    out=[]
    try:
        wb=load_workbook(file_path,read_only=True,data_only=True)
        ws=wb.active
        seen=set()
        for row in ws.iter_rows():
            if not row:
                continue
            uid=normalize_uid(row[0].value)
            if uid and uid not in seen:
                seen.add(uid)
                out.append(uid)
        wb.close()
    except Exception:
        return []
    return out

def report_info():
    cleanup_expired_report()
    r=c.execute("SELECT * FROM reports ORDER BY id DESC LIMIT 1").fetchone()
    if not r:
        return None
    return r

def active_reports():
    cleanup_expired_report()
    return c.execute("SELECT * FROM reports WHERE expires_at>? ORDER BY id DESC", (time.time(),)).fetchall()

async def process_report(bot, report_id, green_uids, category_name=""):
    # Process only pending uploads belonging to the selected report category.
    # A UID is globally payable only once.
    if category_name:
        rows=c.execute(
            "SELECT * FROM submissions WHERE status='pending' AND category_name=? ORDER BY id",
            (category_name,)
        ).fetchall()
    else:
        rows=c.execute("SELECT * FROM submissions WHERE status='pending' ORDER BY id").fetchall()
    user_files={}
    for s in rows:
        user_files.setdefault(s["user_id"],[]).append(s)

    results={}
    total_reward=0.0
    processed_users=0

    for uid_user, submissions in user_files.items():
        seen_user=set()
        paid_count=0
        reward=0.0
        matched_by_submission={}
        for s in submissions:
            uids=extract_submission_uids(s["file_path"])
            if not uids:
                continue
            for uid in uids:
                if uid in seen_user:
                    continue
                seen_user.add(uid)
                if uid not in green_uids:
                    continue
                already=c.execute("SELECT 1 FROM paid_uids WHERE uid=?",(uid,)).fetchone()
                if already:
                    continue
                rate=float(s["price"] or 0)
                c.execute(
                    "INSERT INTO paid_uids(uid,user_id,submission_id,report_id,amount,paid_at) VALUES(?,?,?,?,?,?)",
                    (uid,uid_user,s["id"],report_id,rate,now())
                )
                paid_count += 1
                reward += rate
                matched_by_submission[s["id"]]=matched_by_submission.get(s["id"],0)+1

        if paid_count:
            processed_users += 1
            c.execute(
                "UPDATE users SET balance=balance+?,total_income=total_income+?,today_income=today_income+? WHERE user_id=?",
                (reward,reward,reward,uid_user)
            )
            total_reward += reward
            results[uid_user]=(paid_count,reward,matched_by_submission)
            # Mark submissions with no remaining unprocessed numeric UIDs as completed.
            for s in submissions:
                all_uids=extract_submission_uids(s["file_path"])
                remaining=False
                for uid in all_uids:
                    if not c.execute("SELECT 1 FROM paid_uids WHERE uid=?",(uid,)).fetchone():
                        remaining=True
                        break
                if all_uids and not remaining:
                    c.execute("UPDATE submissions SET status='approved' WHERE id=?",(s["id"],))
        db.commit()

    c.execute("UPDATE reports SET green_count=?,processed_users=?,total_reward=? WHERE id=?",
              (len(green_uids),processed_users,total_reward,report_id))
    db.commit()
    return results, total_reward, processed_users


# --- Admin-configurable Disable Notice + category-specific Work Update emoji ---
DEFAULT_DISABLE_NOTICE = (
    "╭━━━〔 {{PE:5809671801667589890}} SUBMISSION NOTICE 〕━━━╮\n"
    "│\n"
    "│ {{PE:6237640360428837389}} 𝐒𝐔𝐁𝐌𝐈𝐒𝐒𝐈𝐎𝐍 𝐂𝐋𝐎𝐒𝐄𝐃\n"
    "│\n"
    "│ {{PE:6298747815086524010}} এডমিন বর্তমানে ফাইল জমা\n"
    "│    নেওয়া সাময়িকভাবে বন্ধ রেখেছেন।\n"
    "│\n"
    "│ ─────────────────────────\n"
    "│\n"
    "│ 𝐒𝐔𝐁𝐌𝐈𝐒𝐒𝐈𝐎𝐍 𝐒𝐂𝐇𝐄𝐃𝐔𝐋𝐄\n"
    "│ {{PE:5291824687096027834}} বাংলাদেশ সময়\n"
    "│ {{PE:4967525849902350972}} 07:00 PM — 09:30 PM\n"
    "│\n"
    "╰━━━━━━━━━━━━━━━━━━━━━━━━━━╯"
)

WORK_UPDATE_HEADER_EMOJI = "5870692618244984670"
WORK_UPDATE_CATEGORY_FALLBACK_EMOJI = "6298790803414189709"
WORK_UPDATE_CLOSED_EMOJI = "5933728045067145810"


def category_custom_emoji_id(category_name):
    """Pick the Premium/Custom Emoji already defined in the bot for a category."""
    name = (category_name or "").strip().lower()
    if "facebook" in name:
        return PREMIUM_EMOJI_IDS.get("facebook") or WORK_UPDATE_CLOSED_EMOJI
    if "instagram" in name:
        return PREMIUM_EMOJI_IDS.get("instagram") or WORK_UPDATE_CLOSED_EMOJI
    # Reuse any other known category mapping in PREMIUM_EMOJI_IDS when present.
    for key, eid in PREMIUM_EMOJI_IDS.items():
        if key in {"facebook", "instagram"}:
            continue
        if key and key in name:
            return eid
    return WORK_UPDATE_CATEGORY_FALLBACK_EMOJI


def _serialize_entities(entities):
    if not entities:
        return []
    out=[]
    for e in entities:
        try:
            d=e.model_dump(exclude_none=True)
        except Exception:
            try:
                d=dict(e)
            except Exception:
                d={}
        # Telegram entity models may contain enum values; normalize them to strings.
        for k,v in list(d.items()):
            if hasattr(v, "value"):
                d[k]=v.value
        out.append(d)
    return out


def _deserialize_entities(raw):
    if not raw:
        return None
    try:
        data=json.loads(raw) if isinstance(raw,str) else raw
        return [MessageEntity(**x) for x in data]
    except Exception as e:
        print(f"disable notice entity restore error: {e}")
        return None


def get_disable_notice_payload():
    text=setting("disable_notice_text", "")
    if not text:
        text=DEFAULT_DISABLE_NOTICE
    entities=_deserialize_entities(setting("disable_notice_entities", ""))
    # Stored entities belong to the exact stored text and must bypass premiumize.
    if entities:
        return text, entities
    return premiumize(text)


async def send_disable_notice(bot, chat_id, reply_markup=None):
    text, entities=get_disable_notice_payload()
    kwargs={"reply_markup": reply_markup} if reply_markup is not None else {}
    if entities:
        kwargs["entities"]=entities
        kwargs["parse_mode"]=None
    return await bot.send_message(chat_id, text, **kwargs)


async def notify_work_disabled(bot, category_name):
    icon=category_custom_emoji_id(category_name)
    text=(
        f"{{{{PE:{WORK_UPDATE_HEADER_EMOJI}}}}} WORK UPDATE\n\n"
        f"{{{{PE:{icon}}}}} {category_name} কাজটি বর্তমানে বন্ধ করা হয়েছে।\n\n"
        f"{{{{PE:{WORK_UPDATE_CATEGORY_FALLBACK_EMOJI}}}}} এই Category-তে এখন নতুন ফাইল জমা নেওয়া হচ্ছে না।\n\n"
        f"{{{{PE:5409260938488458240}}}} অন্য কোনো active Work থাকলে সেটি ব্যবহার করতে পারবেন।"
    )
    sent=failed=0
    users=c.execute("SELECT user_id FROM users WHERE COALESCE(blocked,0)=0").fetchall()
    for row in users:
        try:
            await sendmsg(bot, row["user_id"], text, reply_markup=um(row["user_id"]))
            sent += 1
        except Exception as e:
            failed += 1
            print(f"work disable notification failed for {row['user_id']}: {e}")
    return sent, failed

class MaintenanceMiddleware(BaseMiddleware):
    async def __call__(self, handler, event, data):
        try:
            uid = event.from_user.id if event.from_user else None
        except Exception:
            uid = None
        if uid and uid != ADMIN_ID:
            try:
                enabled = setting("maintenance_mode", "0") == "1"
            except Exception:
                enabled = False
            if enabled:
                text = (
                    "🔧 <b>Bot Maintenance চলছে</b>\n\n"
                    "বর্তমানে Bot-এ রক্ষণাবেক্ষণের কাজ চলছে।\n"
                    "দয়া করে কিছুক্ষণ পরে আবার চেষ্টা করুন।"
                )
                try:
                    if hasattr(event, "answer") and event.__class__.__name__ == "CallbackQuery":
                        await event.answer("🔧 Maintenance চলছে।", show_alert=True)
                    elif hasattr(event, "answer"):
                        await event.answer(text)
                except Exception:
                    pass
                return None
        return await handler(event, data)


def now(): return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
def setting(key, default=""):
    r=c.execute("SELECT value FROM bot_settings WHERE key=?",(key,)).fetchone()
    return r["value"] if r else default

def save_setting(key,value):
    c.execute("INSERT INTO bot_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",(key,value))
    db.commit()
def user(u):
    day=datetime.now(timezone.utc).date().isoformat()
    r=c.execute("SELECT * FROM users WHERE user_id=?",(u.id,)).fetchone()
    full_name = " ".join(x for x in [u.first_name or "", u.last_name or ""] if x).strip()
    if not r: c.execute("INSERT INTO users(user_id,username,full_name,today_date,created_at) VALUES(?,?,?,?,?)",(u.id,u.username or "",full_name,day,now()))
    elif r["today_date"]!=day: c.execute("UPDATE users SET today_income=0,today_date=?,username=?,full_name=? WHERE user_id=?",(day,u.username or "",full_name,u.id))
    else: c.execute("UPDATE users SET username=?,full_name=? WHERE user_id=?",(u.username or "",full_name,u.id))
    db.commit()

def um(uid):
    # ReplyKeyboard buttons use plain text only so Telegram callback/text
    # matching remains reliable. Custom Premium Emojis are used in messages.
    k=[
        [kbbtn(text="Account")],
        [kbbtn(text="Report"),kbbtn(text="Sell Account")],
        [kbbtn(text="Withdrawal"),kbbtn(text="Price List")],
        [kbbtn(text="Support")]
    ]
    custom=c.execute("SELECT name FROM custom_buttons WHERE active=1 ORDER BY id").fetchall()
    for i in range(0, len(custom), 2):
        k.append([kbbtn(text=x["name"]) for x in custom[i:i+2]])
    if uid==ADMIN_ID:
        k.append([kbbtn(text="Admin Panel")])
    return ReplyKeyboardMarkup(keyboard=k,resize_keyboard=True,)

def pending_admin_kb():
    # Intentionally uses only standard Telegram ReplyKeyboardButton fields.
    # This screen must remain usable even if a deployed aiogram version does
    # not support custom keyboard icon/style fields.
    return ReplyKeyboardMarkup(
        keyboard=[
            [_KeyboardButton(text="User Management"), _KeyboardButton(text="Work Management")],
            [_KeyboardButton(text="Withdrawal Management")],
            [_KeyboardButton(text="Force Join"), _KeyboardButton(text="Pending Files")],
            [_KeyboardButton(text="Upload Report File"), _KeyboardButton(text="Generate Main Excel")],
            [_KeyboardButton(text="Set Payment Channel")],
            [_KeyboardButton(text="Set Receiver Channel")],
            [_KeyboardButton(text="Broadcast"), _KeyboardButton(text="New Button")],
            [_KeyboardButton(text="Database"), _KeyboardButton(text="Maintenance")],
            [_KeyboardButton(text="Support Management")],
            [_KeyboardButton(text="User Menu")],
        ],
        resize_keyboard=True,
    )

def am():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="User Management"),kbbtn(text="Work Management")],
            [kbbtn(text="Withdrawal Management")],
            [kbbtn(text="Force Join"),kbbtn(text="Pending Files")],
            [kbbtn(text="📊 Upload Report File"),kbbtn(text="Generate Main Excel")],
            [kbbtn(text="Set Payment Channel")],
            [kbbtn(text="Set Receiver Channel")],
            [kbbtn(text="Broadcast"),kbbtn(text="New Button")],
            [kbbtn(text="Database"), kbbtn(text="Maintenance")],
            [kbbtn(text="Support Management")],
            [kbbtn(text="User Menu")]
        ],
        resize_keyboard=True,
    )
def cats():
    r=c.execute("SELECT * FROM categories WHERE active=1 ORDER BY id").fetchall()
    rows=[]
    for x in r:
        icon = category_custom_emoji_id(x["name"])
        rows.append([kbbtn(text=f"{x['name']} — {x['price']:.2f}", icon_custom_emoji_id=icon)])
    rows.append([kbbtn(text="Cancel")])
    return ReplyKeyboardMarkup(keyboard=rows,resize_keyboard=True),r
def report_categories_kb():
    rows=c.execute("SELECT * FROM categories WHERE active=1 ORDER BY id").fetchall()
    return ReplyKeyboardMarkup(
        keyboard=[[kbbtn(text=x["name"], icon_custom_emoji_id=category_custom_emoji_id(x["name"]))] for x in rows] + [[kbbtn(text="Cancel")]],
        resize_keyboard=True
    ), rows

def wcats():
    r=c.execute("SELECT * FROM withdrawal_categories WHERE active=1 ORDER BY id").fetchall()
    # Reply keyboard button text stays plain for reliable matching;
    # the Bot API button fields carry the Custom Emoji icon and style.
    rows=[]
    for x in r:
        name = x["name"].strip()
        low = name.lower()
        if "বিকাশ" in name or "bkash" in low:
            icon = PE["bkash"]
        elif "নগদ" in name or "nagad" in low or "nogod" in low:
            icon = PE["nagad"]
        else:
            icon = None
        kwargs = {"icon_custom_emoji_id": icon} if icon else {}
        rows.append([kbbtn(text=name, **kwargs)])
    rows.append([kbbtn(text="Cancel")])
    return ReplyKeyboardMarkup(keyboard=rows,resize_keyboard=True), r

state={}


async def force_join_keyboard(bot, uid):
    rows=c.execute("SELECT * FROM force_join_channels WHERE active=1 ORDER BY id").fetchall()
    missing=[]
    for x in rows:
        try:
            member=await bot.get_chat_member(chat_id=x["chat_id"], user_id=uid)
            # Telegram can return "restricted" for a member who is still in the
            # channel. Only treat restricted users as missing when is_member=False.
            joined = member.status in ("member", "administrator", "creator")
            if member.status == "restricted":
                joined = bool(getattr(member, "is_member", False))
            if not joined:
                missing.append(x)
        except Exception as e:
            # If the bot cannot inspect a required channel, keep it visible so
            # the requirement cannot be silently bypassed.
            print(f"Force Join check error for {x['chat_id']}: {e}")
            missing.append(x)

    if not missing:
        return None

    buttons=[]
    for x in missing:
        link=(x["invite_link"] or "").strip()
        if link:
            buttons.append([ibtn(text=f"Join {x['channel_name']}", url=link, style="primary")])
    buttons.append([ibtn(text="✓ Check Joined", callback_data="check_force_join", style="success")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def force_join_prompt_text():
    return (
        "🔐 Channel Membership Required\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "বট ব্যবহার করার আগে নিচের সবগুলো চ্যানেলে Join করতে হবে।\n\n"
        "① প্রতিটি Join বাটনে চাপ দিন\n"
        "② চ্যানেলে Join সম্পন্ন করুন\n"
        "③ সবশেষে ✓ Check Joined চাপুন\n\n"
        "সবগুলো চ্যানেলে Join হয়ে গেলে আপনি সঙ্গে সঙ্গে বট ব্যবহার করতে পারবেন।"
    )

async def ensure_joined(m):
    if m.from_user.id==ADMIN_ID:
        return True
    kb=await force_join_keyboard(m.bot,m.from_user.id)
    if kb:
        await ans(m, force_join_prompt_text(), reply_markup=kb)
        return False
    return True

async def force_join_callback(q):
    if q.from_user.id==ADMIN_ID:
        await q.answer("Admin bypass enabled.")
        return
    kb=await force_join_keyboard(q.bot,q.from_user.id)
    if kb:
        await q.answer("এখনও সব required channel-এ Join করা হয়নি।", show_alert=True)
        try:
            await q.message.edit_text(force_join_prompt_text(), reply_markup=kb)
        except Exception:
            try:
                await q.message.edit_reply_markup(reply_markup=kb)
            except Exception:
                pass
    else:
        await q.answer("✓ সবগুলো channel Join করা হয়েছে!")
        try:
            await q.message.delete()
        except Exception:
            pass
        await send_same_start_welcome(q.message, q.from_user)

def force_join_admin_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Add Force Join Channel")],
            [kbbtn(text="Force Join Channel List")],
            [kbbtn(text="Remove Force Join Channel")],
            [kbbtn(text="Back to Admin Panel")]
        ],resize_keyboard=True,)

async def force_join_management(m):
    if m.from_user.id!=ADMIN_ID:return
    state.pop(ADMIN_ID,None)
    await ans(m, "Force Join Management",reply_markup=force_join_admin_kb())

async def force_join_admin_action(m):
    if m.from_user.id!=ADMIN_ID:return
    t=m.text.strip()
    if t=="Back to Admin Panel":
        state.pop(ADMIN_ID,None)
        return await ans(m, "Admin Panel",reply_markup=withdrawal_admin_kb())

    if t=="Add Force Join Channel":
        state[ADMIN_ID]={"action":"fj_chat_id"}
        return await ans(m, "ধাপ ১/৩\nChannel ID পাঠান।\nউদাহরণ: -1001234567890",reply_markup=force_join_admin_kb())

    if t=="Force Join Channel List":
        r=c.execute("SELECT * FROM force_join_channels ORDER BY id").fetchall()
        if not r:return await ans(m, "কোনো Force Join channel নেই.",reply_markup=force_join_admin_kb())
        text="Force Join Channels\\n\\n"
        for x in r:
            text += f"ID: {x['id']}\\nChannel: {x['channel_name']}\\nChat ID: {x['chat_id']}\\nStatus: {'ON' if x['active'] else 'OFF'}\\n\\n"
        return await ans(m, text,reply_markup=force_join_admin_kb())

    if t=="Remove Force Join Channel":
        state[ADMIN_ID]={"action":"fj_remove"}
        return await ans(m, "যে channel remove করবেন তার তালিকার ID পাঠান।",reply_markup=force_join_admin_kb())

    a=state.get(ADMIN_ID,{}).get("action")
    try:
        if a=="add_button_name":
            name=m.text.strip()
            if not name or len(name)>50: raise ValueError()
            state[ADMIN_ID]={"action":"add_button_target","name":name}
            return await ans(m, "এখন Button-এর Link অথবা User ID পাঠান।\nউদাহরণ: https://t.me/example অথবা 123456789")
        if a=="add_button_target":
            s=state[ADMIN_ID]; target=m.text.strip()
            if not target: raise ValueError()
            c.execute("INSERT INTO custom_buttons(name,target,active,created_at) VALUES(?,?,1,?) ON CONFLICT(name) DO UPDATE SET target=excluded.target,active=1",(s["name"],target,now()))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Button '{s['name']}' সফলভাবে যোগ/আপডেট হয়েছে।",reply_markup=withdrawal_admin_kb())
        if a=="receiver_channel":
            target=m.text.strip()
            if target.upper()=="OFF":
                save_setting("receiver_channel","")
                state.pop(ADMIN_ID,None)
                return await ans(m, "✅ Receiver Channel বন্ধ করা হয়েছে। এখন file Admin Panel-এ যাবে.",reply_markup=am())
            target=normalize_receiver_target(target)
            if not target: raise ValueError()
            save_setting("receiver_channel",target)
            state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Receiver Channel saved: {target}\n\nBot-কে ওই channel-এ post করার permission দিন.",reply_markup=am())
        if a=="fj_chat_id":
            chat_id=m.text.strip()
            state[ADMIN_ID]={"action":"fj_name","chat_id":chat_id}
            return await ans(m, "ধাপ ২/৩\nChannel-এর নাম পাঠান।",reply_markup=force_join_admin_kb())

        if a=="fj_name":
            s=state[ADMIN_ID]
            s["channel_name"]=m.text.strip()
            s["action"]="fj_link"
            state[ADMIN_ID]=s
            return await ans(m, "ধাপ ৩/৩\nChannel-এর Join/Invite Link পাঠান।",reply_markup=force_join_admin_kb())

        if a=="fj_link":
            s=state[ADMIN_ID]
            c.execute(
                "INSERT INTO force_join_channels(chat_id,channel_name,invite_link,active,created_at) VALUES(?,?,?,?,?)",
                (s["chat_id"],s["channel_name"],m.text.strip(),1,now())
            )
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Force Join channel সফলভাবে যোগ হয়েছে।",reply_markup=force_join_admin_kb())

        if a=="fj_remove":
            cid=int(m.text.strip())
            c.execute("UPDATE force_join_channels SET active=0 WHERE id=?",(cid,))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Force Join channel remove হয়েছে।",reply_markup=force_join_admin_kb())
    except:
        await ans(m, "তথ্যটি সঠিক নয়। আবার চেষ্টা করুন।",reply_markup=force_join_admin_kb())


def work_management_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [kbbtn(text="Add Work"), kbbtn(text="Edit Work")],
        [kbbtn(text="Disable Work"), kbbtn(text="Enable Work")],
        [kbbtn(text="Disable Notice")],
        [kbbtn(text="Work List")],
        [kbbtn(text="Back to Admin Panel")],
    ], resize_keyboard=True)

async def work_menu_action(m):
    if m.from_user.id != ADMIN_ID: return
    t = clean_button_text((m.text or "").strip())
    st = state.get(ADMIN_ID, {})
    a = st.get("action")
    try:
        if t == "Back to Admin Panel" and (not a or a == "work_menu"):
            state.pop(ADMIN_ID, None); return await ans(m, "Admin Panel", reply_markup=am())
        if t == "Work List" and (not a or a == "work_menu"):
            rows=c.execute("SELECT id,name,price,active FROM categories ORDER BY id").fetchall()
            if not rows: text="📂 Work List\n\nকোনো Work এখনো যোগ করা হয়নি।"
            else:
                text="📂 Work List\n\n"+"\n".join(f"🆔 ID: {x['id']}\n📁 {x['name']}\n💰 Price: {float(x['price'] or 0):.2f}/account\n📌 Status: {'🟢 ACTIVE' if x['active'] else '🔴 DISABLED'}\n" for x in rows)
            return await ans(m,text,reply_markup=work_management_kb())
        if t == "Add Work" and (not a or a == "work_menu"):
            state[ADMIN_ID]={"action":"work_name"}; return await ans(m,"➕ Add Work\n\nWork/Category-এর নাম পাঠান।")
        if t == "Edit Work" and (not a or a == "work_menu"):
            rows=c.execute("SELECT id,name,price,active FROM categories ORDER BY id").fetchall()
            if not rows: return await ans(m,"❌ Edit করার মতো কোনো Work নেই।",reply_markup=work_management_kb())
            state[ADMIN_ID]={"action":"work_edit_id"}
            return await ans(m,"✏️ Edit Work\n\n"+"\n".join(f"ID {x['id']} — {x['name']} — {float(x['price'] or 0):.2f} — {'ON' if x['active'] else 'OFF'}" for x in rows)+"\n\nWork ID পাঠান।")
        if t in ("Disable Work","Enable Work") and (not a or a == "work_menu"):
            rows=c.execute("SELECT id,name,active FROM categories ORDER BY id").fetchall()
            if not rows: return await ans(m,"❌ কোনো Work নেই।",reply_markup=work_management_kb())
            toggle=0 if t=="Disable Work" else 1
            state[ADMIN_ID]={"action":"work_toggle_id","toggle":toggle}
            return await ans(m,("🔴 Disable Work" if not toggle else "🟢 Enable Work")+"\n\n"+"\n".join(f"ID {x['id']} — {x['name']} — {'ON' if x['active'] else 'OFF'}" for x in rows)+"\n\nWork ID পাঠান।")
        if t == "Disable Notice" and (not a or a == "work_menu"):
            state[ADMIN_ID]={"action":"disable_notice"}
            return await ans(m,"🔕 Disable Notice\n\nকোনো Active Work না থাকলে User-এর কাছে যে message দেখাবে সেটি পাঠান।\nPremium/Custom Emoji ও formatting সংরক্ষণ করা হবে।",reply_markup=ReplyKeyboardRemove())
        if a == "work_name":
            name=(m.text or "").strip()
            if name.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            if not name or len(name)>120: return await ans(m,"❌ Work name 1–120 characters দিন।")
            dup=c.execute("SELECT id FROM categories WHERE lower(name)=lower(?)",(name,)).fetchone()
            if dup: state.pop(ADMIN_ID,None); return await ans(m,f"❌ এই Work আগে থেকেই আছে। ID: {dup['id']}",reply_markup=work_management_kb())
            state[ADMIN_ID]={"action":"work_price","name":name}; return await ans(m,"💰 প্রতি Account-এর Price পাঠান।")
        if a == "work_price":
            raw=(m.text or "").strip()
            if raw.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            price=float(raw)
            if price<0 or price>1000000: raise ValueError
            c.execute("INSERT INTO categories(name,price,active) VALUES(?,?,1)",(st["name"],price)); db.commit(); wid=c.lastrowid; state.pop(ADMIN_ID,None)
            return await ans(m,f"✅ Work যোগ হয়েছে।\n🆔 ID: {wid}\n📁 {st['name']}\n💰 {price:.2f}/account",reply_markup=work_management_kb())
        if a == "work_edit_id":
            raw=(m.text or "").strip()
            if raw.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            cid=int(raw); cat=c.execute("SELECT * FROM categories WHERE id=?",(cid,)).fetchone()
            if not cat: return await ans(m,"❌ এই ID-এর কোনো Work নেই।")
            state[ADMIN_ID]={"action":"work_edit_name","id":cid}; return await ans(m,f"নতুন Work Name পাঠান।\nবর্তমান: {cat['name']}")
        if a == "work_edit_name":
            name=(m.text or "").strip()
            if name.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            if not name or len(name)>120: return await ans(m,"❌ Work name 1–120 characters দিন।")
            dup=c.execute("SELECT id FROM categories WHERE lower(name)=lower(?) AND id<>?",(name,st["id"])).fetchone()
            if dup: return await ans(m,"❌ এই নামে আরেকটি Work আছে।")
            st["name"]=name; st["action"]="work_edit_price"; state[ADMIN_ID]=st; return await ans(m,"নতুন Price পাঠান।")
        if a == "work_edit_price":
            raw=(m.text or "").strip()
            if raw.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            price=float(raw)
            if price<0 or price>1000000: raise ValueError
            cur=c.execute("UPDATE categories SET name=?,price=? WHERE id=?",(st["name"],price,int(st["id"])))
            if cur.rowcount!=1: raise ValueError
            db.commit(); cid=st["id"]; name=st["name"]; state.pop(ADMIN_ID,None)
            return await ans(m,f"✅ Work update হয়েছে।\n🆔 ID: {cid}\n📁 {name}\n💰 {price:.2f}/account",reply_markup=work_management_kb())
        if a == "work_toggle_id":
            raw=(m.text or "").strip()
            if raw.lower()=="cancel": state.pop(ADMIN_ID,None); return await ans(m,"❌ বাতিল।",reply_markup=work_management_kb())
            cid=int(raw); toggle=int(st["toggle"]); cat=c.execute("SELECT * FROM categories WHERE id=?",(cid,)).fetchone()
            if not cat: return await ans(m,"❌ এই ID-এর কোনো Work নেই।")
            was_active=int(cat["active"] or 0)==1; c.execute("UPDATE categories SET active=? WHERE id=?",(toggle,cid)); db.commit(); state.pop(ADMIN_ID,None)
            if toggle==0 and was_active:
                sent,failed=await notify_work_disabled(m.bot,cat["name"])
                return await ans(m,f"🔕 {cat['name']} Work বন্ধ করা হয়েছে।\n\nNotification Sent: {sent} | Failed: {failed}",reply_markup=work_management_kb())
            return await ans(m,f"✅ {cat['name']} এখন {'🟢 enabled' if toggle else '🔴 disabled'}।",reply_markup=work_management_kb())
        if a == "disable_notice":
            text=m.text if m.text is not None else m.caption; entities=m.entities if m.text is not None else m.caption_entities
            if not text: return await ans(m,"❌ একটি text message পাঠান।")
            save_setting("disable_notice_text",text); save_setting("disable_notice_entities",json.dumps(_serialize_entities(entities),ensure_ascii=False)); state.pop(ADMIN_ID,None)
            return await ans(m,"✅ Disable Notice database-এ save হয়েছে।",reply_markup=work_management_kb())
    except (ValueError,TypeError):
        return await ans(m,"❌ তথ্যটি সঠিক নয়। ID/Price সঠিকভাবে দিন।")
    except sqlite3.Error as e:
        db.rollback(); print("[work-management] DB error:",e); return await ans(m,"❌ Database error হয়েছে। আগের data নিরাপদ আছে।",reply_markup=work_management_kb())
    except Exception as e:
        print("[work-management] error:",e); return await ans(m,"❌ Work Management error হয়েছে। আবার চেষ্টা করুন।",reply_markup=work_management_kb())

def normalize_receiver_target(value):
    v=(value or "").strip()
    if not v:
        return ""
    if re.match(r"^-100\d+$",v) or v.isdigit():
        return v
    m=re.match(r"^(?:https?://)?(?:www\.)?t\.me/([A-Za-z0-9_]+)(?:/.*)?$",v)
    if m:
        return "@"+m.group(1)
    if v.startswith("@"):
        return v
    return v

async def handle_custom_button(m):
    row=c.execute("SELECT * FROM custom_buttons WHERE name=? AND active=1",(clean_button_text(m.text),)).fetchone()
    if not row:
        return False
    target=row["target"].strip()
    if target.isdigit():
        kb=InlineKeyboardMarkup(inline_keyboard=[[ibtn(text="👤 Open User",url=f"tg://user?id={target}")]])
        await ans(m, f"User ID: {target}",reply_markup=kb)
        return True
    if target.startswith("@"):
        url="https://t.me/"+target[1:]
    elif re.match(r"^https?://",target):
        url=target
    elif re.match(r"^(?:www\.)?t\.me/",target):
        url="https://"+target
    else:
        url="https://"+target
    kb=InlineKeyboardMarkup(inline_keyboard=[[ibtn(text=f"🔗 {row['name']}",url=url)]])
    await ans(m, f"{row['name']}:",reply_markup=kb)
    return True

async def send_welcome(bot, user):
    name = (user.first_name or user.username or "User").strip()
    text = f"🥰 স্বাগতম, {name}!\n💎কাজ শুরু করতে নিচের অপশনগুলো ব্যবহার করুন 🔽"
    try:
        await sendmsg(bot, user.id, text)
    except Exception:
        pass

async def userpanel(m):
    if m.from_user.id!=ADMIN_ID:
        return await ans(m, "User Menu",reply_markup=um(m.from_user.id))
    state.pop(ADMIN_ID,None)
    await ans(m, "User Menu",reply_markup=um(ADMIN_ID))


async def send_same_start_welcome(m, user):
        name = (user.first_name or "User").strip()
        welcome_text = f"🥰 স্বাগতম, {name}!\n💎কাজ শুরু করতে নিচের অপশনগুলো ব্যবহার করুন 🔽"
        await ans(m, welcome_text, reply_markup=um(user.id))

async def start(m):
    # /start must always clear stale multi-step states.
    state.pop(m.from_user.id, None)
    if m.from_user.id!=ADMIN_ID:
        r=c.execute("SELECT blocked FROM users WHERE user_id=?",(m.from_user.id,)).fetchone()
        if r and r["blocked"]:
            return await ans(m, "Your account is blocked.")
    user(m.from_user)
    if not await ensure_joined(m): return
    await send_same_start_welcome(m, m.from_user)

async def account(m):
    if not await ensure_joined(m): return
    state.pop(m.from_user.id, None)
    user(m.from_user); r=c.execute("SELECT * FROM users WHERE user_id=?",(m.from_user.id,)).fetchone()
    await ans(m, f"Account\n\nUser ID: {r['user_id']}\nBalance: {r['balance']:.2f}\nToday's Income: {r['today_income']:.2f}\nTotal Income: {r['total_income']:.2f}",reply_markup=um(m.from_user.id))
async def send_report_file(m, r, user_id=None):
    expires=max(0,int(float(r["expires_at"])-time.time()))
    hrs=expires//3600; mins=(expires%3600)//60
    text=(
        "📊 Report Available\n\n"
        f"📁 Report: {r['file_name']}\n"
        f"⏳ Report expires in: {hrs}h {mins}m\n\n"
        "রিপোর্ট এসে গেছে এডমিন আপনার রিপোর্ট চেক করে ব্যালেন্স এড করে দিবে ধৈর্য ধরুন।\n\n"
        "রিপোর্টে ভুল আসলে আপনি নিজেও ফাইলটি চেক করে দেখতে পারেন যদি কোন সমস্যা খুঁজে পান তাহলে সাপোর্টে যোগাযোগ করবেন ধন্যবাদ"
    )
    try:
        raw=Path(r["file_path"]).read_bytes()
        target_id=user_id if user_id is not None else m.from_user.id
        return await senddoc(m.bot, 
            target_id,
            BufferedInputFile(raw,filename=r["file_name"]),
            caption=text
        )
    except Exception:
        return await ans(m, text,reply_markup=um(m.from_user.id))

async def report_category_callback(q):
    # Admin can also test the User Menu/report flow. Future Reports remains
    # available from Admin Panel for viewing the full active report list.
    cleanup_expired_report()
    try:
        rid=int(q.data.split(":")[1])
    except Exception:
        return await q.answer("Invalid report.", show_alert=True)
    r=c.execute("SELECT * FROM reports WHERE id=? AND expires_at>?",(rid,time.time())).fetchone()
    if not r:
        return await q.answer("এই রিপোর্টটি আর active নেই।", show_alert=True)
    await q.answer()
    # Edit the selector away so the user receives only the requested file message.
    try:
        await q.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await send_report_file(q.message, r, q.from_user.id)

async def report_section(m):
    if not await ensure_joined(m): return
    state.pop(m.from_user.id, None)
    reports=active_reports()
    if not reports:
        rows=c.execute(
            "SELECT category_name,SUM(rows_count) n,SUM(estimated_profit) v "
            "FROM submissions WHERE user_id=? AND status='pending' GROUP BY category_name",
            (m.from_user.id,)
        ).fetchall()
        pending_text=(
            "👤 Your Accounts: 0\n💰 Estimated Profit: 0.00"
            if not rows else
            "".join(f"👤 {x['category_name']}: {int(x['n'] or 0)} accounts\n💰 Estimated Profit: {float(x['v'] or 0):.2f}\n" for x in rows)
        )
        return await ans(m, 
            "📊 Report\n\n"
            "⏳ Coming Soon...\n"
            "বর্তমানে কোনো রিপোর্ট আপলোড করা হয়নি।\n\n"
            + pending_text,
            reply_markup=um(m.from_user.id)
        )

    if len(reports)==1:
        return await send_report_file(m, reports[0])

    kb=[
        [ibtn(
            text=f"📁 {r['category_name'] or 'Report'}",
            callback_data=f"report:{r['id']}"
        )]
        for r in reports
    ]
    return await ans(m, 
        "📊 Report Available\n\n"
        "একাধিক ক্যাটাগরির রিপোর্ট active আছে। যে রিপোর্টটি দেখতে চান সেটি নির্বাচন করুন:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )

async def sell(m):
    if not await ensure_joined(m): return
    k,r=cats()
    if not r:
        return await send_disable_notice(m.bot, m.from_user.id, reply_markup=um(m.from_user.id))
    state[m.from_user.id]={"cat":True}; await ans(m, "কাজ নির্বাচন করুন। তারপর XLSX file পাঠান.",reply_markup=k)
async def prices(m):
    if not await ensure_joined(m): return
    state.pop(m.from_user.id, None)
    r=c.execute("SELECT name,price FROM categories WHERE active=1").fetchall()
    await ans(m, "Price List\n\n"+("কোনো price list নেই." if not r else "\n".join(f"{x['name']} — {x['price']:.2f}/account" for x in r)),reply_markup=um(m.from_user.id))
async def support(m):
    if not await ensure_joined(m):
        return

    text = """📞 গ্রাহক সেবা কেন্দ্র
━━━━━━━━━━━━━━━━━━━━━━

সম্মানিত মেম্বার,
আপনার যেকোনো সমস্যা বা জিজ্ঞাসার জন্য আমাদের সাপোর্ট টিমের সাথে যোগাযোগ করুন। আমরা দ্রুত সমাধানের চেষ্টা করব।

নোট:
কোন সমস্যার কথা জানাতে দ্বিধাবোধ করবেন না.!
আমাদের সাথে থাকার জন্য ধন্যবাদ।"""

    def normalize_support_url(raw):
        url = (raw or "").strip()
        if not url:
            return ""
        if url.isdigit():
            return "tg://user?id=" + url
        if url.startswith("@"):
            return "https://t.me/" + url[1:]
        if re.match(r"^tg://", url, re.I):
            return url
        if re.match(r"^https?://", url, re.I):
            return url
        if re.match(r"^(?:www\.)?t\.me/[A-Za-z0-9_+/?=&.-]+$", url, re.I):
            return "https://" + url
        return ""

    # Build each support button independently. A bad/old database row must
    # never break the main Support response for the user.
    buttons = []
    main_url = normalize_support_url(setting("support_link"))
    if main_url:
        buttons.append([ibtn(
            "Contact Support",
            url=main_url,
            icon_custom_emoji_id=PE.get("support")
        )])

    rows = c.execute(
        "SELECT id,name,url FROM support_buttons WHERE active=1 ORDER BY id"
    ).fetchall()

    valid_rows = []
    for r in rows:
        url = normalize_support_url(r["url"])
        name = (r["name"] or "Support").strip()
        if url and name:
            valid_rows.append((name, url))

    for i in range(0, len(valid_rows), 2):
        row = []
        for name, url in valid_rows[i:i+2]:
            row.append(ibtn(
                name,
                url=url,
                icon_custom_emoji_id=PE.get("support")
            ))
        if row:
            buttons.append(row)

    # Always answer the Support message. If Telegram rejects a particular
    # custom-emoji/style combination, retry the same support links without
    # optional button metadata instead of leaving the user with no response.
    if buttons:
        try:
            return await ans(m, text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
        except Exception as e:
            print(f"support premium keyboard error: {e}")
            fallback = []
            if main_url:
                fallback.append([_InlineKeyboardButton(text="Contact Support", url=main_url)])
            for i in range(0, len(valid_rows), 2):
                fallback.append([
                    _InlineKeyboardButton(text=name, url=url)
                    for name, url in valid_rows[i:i+2]
                ])
            try:
                return await ans(m, text, reply_markup=InlineKeyboardMarkup(inline_keyboard=fallback))
            except Exception as e2:
                print(f"support fallback keyboard error: {e2}")

    return await ans(
        m,
        text + "\n\nসাপোর্ট লিংক বর্তমানে সেট করা হয়নি।",
        reply_markup=um(m.from_user.id)
    )

async def withdraw(m):
    if not await ensure_joined(m): return
    user(m.from_user)
    k,r=wcats()
    if not r:return await ans(m, "বর্তমানে কোনো withdrawal category active নেই.",reply_markup=um(m.from_user.id))
    state[m.from_user.id]={"wcat":True}; await ans(m, "Withdrawal category নির্বাচন করুন.",reply_markup=k)

async def doc(m):
    if not await ensure_joined(m): return
    user(m.from_user); s=state.get(m.from_user.id,{})
    if "category_id" not in s:
        return await ans(m, "আগে Sell Account চাপুন → Work নির্বাচন করুন → তারপর XLSX file পাঠান.",reply_markup=um(m.from_user.id))
    fn=m.document.file_name or ""
    if not fn.lower().endswith(".xlsx"):
        return await ans(m, "শুধুমাত্র .xlsx file গ্রহণযোগ্য.",reply_markup=um(m.from_user.id))
    cat=c.execute("SELECT * FROM categories WHERE id=? AND active=1",(s["category_id"],)).fetchone()
    if not cat:return await ans(m, "Category আর active নেই.",reply_markup=um(m.from_user.id))
    f=await m.bot.get_file(m.document.file_id); b=io.BytesIO(); await m.bot.download_file(f.file_path,b); raw=b.getvalue()
    # Reject an exact duplicate upload from the same user before saving or forwarding it.
    file_hash=hashlib.sha256(raw).hexdigest()
    dup=c.execute("SELECT id,file_name,status,created_at FROM submissions WHERE user_id=? AND file_hash=? LIMIT 1",(m.from_user.id,file_hash)).fetchone()
    if dup:
        return await ans(m, 
            "⚠️ এই ফাইলটি আপনি আগে জমা দিয়েছেন.\\n\\n"
            "Duplicate XLSX file গ্রহণ করা হবে না। নতুন/ভিন্ন file upload করুন।",
            reply_markup=um(m.from_user.id)
        )
    try:
        wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True); ws=wb.active
        n=count_xlsx_data_rows(ws); wb.close()
    except Exception:
        return await ans(m, "XLSX file পড়া যায়নি.",reply_markup=um(m.from_user.id))
    if n<=0:return await ans(m, "কোনো account/data row পাওয়া যায়নি.",reply_markup=um(m.from_user.id))
    p=FD/f"{m.from_user.id}_{int(datetime.now().timestamp())}_{Path(fn).name}"; p.write_bytes(raw); profit=n*float(cat["price"])
    receiver=setting("receiver_channel").strip()
    if not receiver:
        try: p.unlink()
        except Exception: pass
        return await ans(m, "⚠️ Receiver Channel এখনো সেট করা হয়নি। Admin Panel → Set Receiver Channel থেকে channel সেট করুন, তারপর file upload করুন।",reply_markup=um(m.from_user.id))
    c.execute("INSERT INTO submissions(user_id,username,category_id,category_name,price,rows_count,estimated_profit,file_name,file_path,status,created_at,file_hash,file_id,receiver_message_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(m.from_user.id,m.from_user.username or "",cat["id"],cat["name"],cat["price"],n,profit,fn,str(p),"pending",now(),file_hash,m.document.file_id,0)); sid=c.lastrowid; db.commit(); state.pop(m.from_user.id,None)
    caption=f"New Account File\n\nSubmission ID: {sid}\nUsername: @{m.from_user.username or 'N/A'}\nUser ID: {m.from_user.id}\nWork: {cat['name']}\nAccounts: {n}\nPrice: {cat['price']:.2f}\nEstimated Profit: {profit:.2f}\nFile: {fn}\nTime: {now()}"
    # Receiver channel receives the file. Save the resulting Telegram message id
    # as a second recovery path for Pending Files.
    try:
        receiver_msg=await senddoc(m.bot, receiver,BufferedInputFile(raw,filename=fn),caption=caption)
        c.execute("UPDATE submissions SET receiver_message_id=? WHERE id=?",(receiver_msg.message_id,sid))
        db.commit()
    except Exception:
        c.execute("DELETE FROM submissions WHERE id=?",(sid,)); db.commit()
        try: p.unlink()
        except Exception: pass
        return await ans(m, "⚠️ Receiver Channel-এ file পাঠানো যায়নি। Public @username/link এবং bot-এর channel permission পরীক্ষা করুন।",reply_markup=um(m.from_user.id))
    await ans(m, f"আপনার file সফলভাবে জমা হয়েছে.\n\nWork: {cat['name']}\nAccount: {n}\nPrice: {cat['price']:.2f}/account\nআনুমানিক Profit: {profit:.2f}\n\nAdmin যাচাই করার পর Balance-এ টাকা যোগ হবে.",reply_markup=um(m.from_user.id))

def count_xlsx_data_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if any(v is not None and str(v).strip() for v in vals):
            rows.append(vals)
    if not rows:
        return 0
    first = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
    header_words = {
        "username", "user name", "userid", "user id", "email", "e-mail",
        "password", "pass", "cookie", "cookies", "account", "accounts",
        "mail", "login", "phone", "number", "mobile", "id"
    }
    # Only treat the first row as a header when it clearly looks like one.
    # This prevents a real account row containing a single word such as
    # "username" from being incorrectly removed from the count.
    header_hits = sum(1 for v in first if v in header_words)
    looks_like_header = header_hits >= 2 or (header_hits == 1 and len(first) > 1 and first[0] in {"username", "user name", "email", "e-mail", "account", "accounts", "userid", "user id"})
    return len(rows) - 1 if looks_like_header else len(rows)


def user_mgmt():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="All User List")],
            [kbbtn(text="Block User"),kbbtn(text="Blocked User")],
            [kbbtn(text="Add Balance"),kbbtn(text="Remove Balance")],
            [kbbtn(text="Back to Admin Panel")]
        ], resize_keyboard=True)

async def user_management(m):
    if m.from_user.id != ADMIN_ID: return
    await ans(m, "User Management", reply_markup=user_mgmt())

async def user_mgmt_action(m):
    if m.from_user.id != ADMIN_ID: return
    t=m.text.strip()
    if t=="Back to Admin Panel":
        state.pop(ADMIN_ID,None)
        return await ans(m, "Admin Panel", reply_markup=am())
    if t=="All User List":
        r=c.execute("SELECT * FROM users ORDER BY user_id DESC").fetchall()
        if not r:
            return await ans(m, "কোনো user নেই.", reply_markup=user_mgmt())
        text="All User List\n\n"
        for x in r:
            status="BLOCKED" if x["blocked"] else "ACTIVE"
            text += f"User ID: {x['user_id']}\nUsername: @{x['username'] or 'N/A'}\nBalance: {x['balance']:.2f}\nStatus: {status}\n\n"
        return await ans(m, text, reply_markup=user_mgmt())
    if t=="Block User":
        state[ADMIN_ID]={"action":"block"}
        return await ans(m, "যে User ID block করতে চান শুধু User ID পাঠান.", reply_markup=user_mgmt())
    if t=="Blocked User":
        r=c.execute("SELECT * FROM users WHERE blocked=1 ORDER BY user_id DESC").fetchall()
        if not r:
            return await ans(m, "কোনো blocked user নেই.", reply_markup=user_mgmt())
        text="Blocked User List\n\n"
        for x in r:
            text += f"User ID: {x['user_id']}\nUsername: @{x['username'] or 'N/A'}\nBalance: {x['balance']:.2f}\n\n"
        return await ans(m, text, reply_markup=user_mgmt())
    if t in ("Add Balance","Remove Balance"):
        state[ADMIN_ID]={"action":"add" if t=="Add Balance" else "remove"}
        return await ans(m, "Format: USER_ID AMOUNT", reply_markup=user_mgmt())


def support_admin_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Add Support Button")],
            [kbbtn(text="Delete Support Button")],
            [kbbtn(text="Support List")],
            [kbbtn(text="Back to Admin Panel")]
        ],
        resize_keyboard=True,
    )

def withdrawal_admin_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Add Withdrawal Category"),kbbtn(text="Delete Category")],
            [kbbtn(text="Pending Withdrawal Requests")],
            [kbbtn(text="Back to Admin Panel")]
        ],resize_keyboard=True,)

async def withdrawal_admin_action(m):
    if m.from_user.id!=ADMIN_ID:return
    t=m.text.strip()
    if t=="Back to Admin Panel":
        state.pop(ADMIN_ID,None)
        return await ans(m, "Admin Panel",reply_markup=am())
    if t=="Add Withdrawal Category":
        state[ADMIN_ID]={"action":"w_add_method"}
        return await ans(m, "Payment method নির্বাচন করুন।", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [ibtn(text="বিকাশ", callback_data="waddmethod:bkash", icon_custom_emoji_id=PE.get("bkash")),
             ibtn(text="নগদ", callback_data="waddmethod:nagad", icon_custom_emoji_id=PE.get("nagad"))]
        ]))
    if t=="Delete Category":
        rows=c.execute("SELECT * FROM withdrawal_categories ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "কোনো সেট করা payment category নেই।", reply_markup=withdrawal_admin_kb())
        await ans(
            m,
            "🗑️ Delete Payment Category\\n\\n"
            "যে payment category মুছে ফেলতে চান সেটি নির্বাচন করুন।\\n"
            "মুছে ফেলার পর category-টি আর নতুন withdrawal-এর জন্য ব্যবহার করা যাবে না।",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [ibtn(text=f"Delete: {x['name']}", callback_data=f"wdelete:{x['id']}", style="danger")]
                for x in rows
            ])
        )
        return await ans(m, "নিচের বাটন থেকে category নির্বাচন করুন।", reply_markup=withdrawal_admin_kb())
    if t in {"Pending Withdrawal Requests","Pending Withdrawals"}:
        # Dedicated handlers above own these buttons.
        rows=c.execute("SELECT id,user_id,username,category_name,amount,wallet,created_at FROM withdrawals WHERE LOWER(TRIM(COALESCE(status,'')))='pending' ORDER BY id DESC").fetchall()
        if not rows:
            return await ans(m, "Pending Withdrawal: 0",reply_markup=withdrawal_admin_kb())
        for x in rows:
            kb=InlineKeyboardMarkup(inline_keyboard=[
                [ibtn("Approve", callback_data=f"wapprove:{x['id']}", icon_custom_emoji_id=PE["done"]),
                 ibtn("Reject + Refund", style="danger", callback_data=f"wreject:{x['id']}", icon_custom_emoji_id=PE["cross"])]
            ])
            await ans(m, 
                f"Withdrawal #{x['id']}\\n"
                f"Username: @{x['username'] or 'N/A'}\\n"
                f"User ID: {x['user_id']}\\n"
                f"Category: {x['category_name']}\\n"
                f"Amount: {x['amount']:.2f}\\n"
                f"Wallet: {x['wallet']}\\n"
                f"Time: {x['created_at']}",
                reply_markup=kb
            )
        return await ans(m, "Pending withdrawal list শেষ।",reply_markup=withdrawal_admin_kb())

    a=state.get(ADMIN_ID,{}).get("action")
    try:
        if a=="w_add_min":
            state[ADMIN_ID]={"action":"w_add_max","name":state[ADMIN_ID]["name"],"minimum":float(m.text.strip())}
            return await ans(m, "Maximum withdrawal amount পাঠান। 0 দিলে unlimited হবে.")
        if a=="w_add_max":
            s=state[ADMIN_ID]
            maximum=float(m.text.strip())
            if maximum>0 and maximum<float(s["minimum"]): raise ValueError()
            state[ADMIN_ID]={"action":"w_add_fee","name":s["name"],"minimum":s["minimum"],"maximum":maximum}
            return await ans(m, "উত্তোলন ফি পাঠান। 0 দিলে কোনো fee থাকবে না।")
        if a=="w_add_fee":
            s=state[ADMIN_ID]; fee=float(m.text.strip())
            if fee<0: raise ValueError()
            c.execute("INSERT INTO withdrawal_categories(name,minimum,maximum,fee,active) VALUES(?,?,?,?,1)",(s["name"],s["minimum"],s["maximum"],fee))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Withdrawal category added.\n\nMethod: {s['name']}\nMinimum: {s['minimum']:.2f}\nMaximum: {s['maximum']:.2f}\nFee: {fee:.2f} টাকা",reply_markup=withdrawal_admin_kb())
        if a=="w_edit_name":
            state[ADMIN_ID]={"action":"w_edit_min","id":state[ADMIN_ID]["id"],"name":m.text.strip()}
            return await ans(m, "নতুন Minimum পাঠান।")
        if a=="w_edit_min":
            state[ADMIN_ID]={"action":"w_edit_max","id":state[ADMIN_ID]["id"],"name":state[ADMIN_ID]["name"],"minimum":float(m.text.strip())}
            return await ans(m, "নতুন Maximum পাঠান। 0 = unlimited.")
        if a=="w_edit_max":
            s=state[ADMIN_ID]; maximum=float(m.text.strip())
            if maximum>0 and maximum<float(s["minimum"]): raise ValueError()
            state[ADMIN_ID]={"action":"w_edit_fee","id":s["id"],"name":s["name"],"minimum":s["minimum"],"maximum":maximum}
            return await ans(m, "নতুন উত্তোলন ফি পাঠান। 0 = no fee.")
        if a=="w_edit_fee":
            s=state[ADMIN_ID]; fee=float(m.text.strip())
            if fee<0: raise ValueError()
            c.execute("UPDATE withdrawal_categories SET name=?,minimum=?,maximum=?,fee=? WHERE id=?",(s["name"],s["minimum"],s["maximum"],fee,s["id"]))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "✅ Withdrawal category updated.",reply_markup=withdrawal_admin_kb())
    except Exception:
        return await ans(m, "Invalid amount/format. আবার চেষ্টা করুন.",reply_markup=withdrawal_admin_kb())

async def support_admin_action(m):
    if m.from_user.id!=ADMIN_ID:
        return
    t=clean_button_text(m.text)
    st=state.get(ADMIN_ID,{})
    a=st.get("action")

    if t=="Back to Admin Panel":
        state.pop(ADMIN_ID,None)
        return await ans(m, "Admin Panel",reply_markup=am())

    if t=="Add Support Button":
        state[ADMIN_ID]={"action":"support_btn_name"}
        return await ans(m, "নতুন Support inline button-এর নাম পাঠান।")

    if t=="Support List":
        rows=c.execute("SELECT id,name,url FROM support_buttons WHERE active=1 ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "📋 Support List\n\nকোনো Support button যোগ করা হয়নি।",reply_markup=support_admin_kb())
        text="📋 Support List\n\n" + "\n\n".join(
            f"ID: {r['id']}\nName: {r['name']}\nURL: {r['url']}" for r in rows
        )
        return await ans(m, text,reply_markup=support_admin_kb())

    if t=="Delete Support Button":
        rows=c.execute("SELECT id,name FROM support_buttons WHERE active=1 ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "কোনো Support button নেই।",reply_markup=support_admin_kb())
        state[ADMIN_ID]={"action":"support_btn_delete_id"}
        return await ans(
            m,
            "যে Support button delete করবেন তার ID পাঠান।\n\n" +
            "\n".join(f"{r['id']}. {r['name']}" for r in rows)
        )

    if a=="support_btn_name":
        name=m.text.strip()
        if not name or len(name)>40:
            return await ans(m, "সঠিক button name দিন।")
        state[ADMIN_ID]={"action":"support_btn_url","name":name}
        return await ans(m, "এখন Support button-এর URL পাঠান।")

    if a=="support_btn_url":
        url=m.text.strip()
        if url.isdigit():
            url="tg://user?id="+url
        elif url.startswith("@"):
            url="https://t.me/"+url[1:]
        elif not re.match(r"^(?:https?://|tg://)",url):
            if re.match(r"^(?:www\\.)?t\\.me/[A-Za-z0-9_+/?=&.-]+$",url):
                url="https://"+url
            else:
                return await ans(m, "সঠিক Link অথবা User ID দিন। উদাহরণ: @support অথবা 123456789")
        name=st.get("name","")
        c.execute(
            "INSERT INTO support_buttons(name,url,active,created_at) VALUES(?,?,1,?)",
            (name,url,now())
        )
        db.commit()
        state.pop(ADMIN_ID,None)
        return await ans(m, f"✅ Support button '{name}' যোগ হয়েছে।",reply_markup=support_admin_kb())

    if a=="support_btn_delete_id":
        try:
            bid=int(m.text.strip())
        except:
            return await ans(m, "সঠিক Support button ID দিন।")
        r=c.execute("SELECT * FROM support_buttons WHERE id=? AND active=1",(bid,)).fetchone()
        if not r:
            return await ans(m, "Button পাওয়া যায়নি।",reply_markup=support_admin_kb())
        c.execute("UPDATE support_buttons SET active=0 WHERE id=?",(bid,))
        db.commit()
        state.pop(ADMIN_ID,None)
        return await ans(m, f"🗑 Support button '{r['name']}' delete হয়েছে।",reply_markup=support_admin_kb())

def maintenance_admin_kb():
    enabled = setting("maintenance_mode", "0") == "1"
    status = "ON" if enabled else "OFF"
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Maintenance ON"), kbbtn(text="Maintenance OFF")],
            [kbbtn(text="Back to Admin Panel")],
        ],
        resize_keyboard=True,
    )


async def maintenance_admin_action(m):
    if m.from_user.id != ADMIN_ID:
        return
    t = clean_button_text(m.text)
    if t == "Maintenance":
        return await ans(m, "🔧 Maintenance\n\nবর্তমান Status: " + ("ON" if setting("maintenance_mode", "0") == "1" else "OFF"), reply_markup=maintenance_admin_kb())
    if t == "Maintenance ON":
        save_setting("maintenance_mode", "1")
        return await ans(m, "🔧 Maintenance Mode ON করা হয়েছে।\n\nএখন সাধারণ User Bot ব্যবহার করতে পারবে না।", reply_markup=maintenance_admin_kb())
    if t == "Maintenance OFF":
        save_setting("maintenance_mode", "0")
        return await ans(m, "✅ Maintenance Mode OFF করা হয়েছে।\n\nUser আবার Bot ব্যবহার করতে পারবে।", reply_markup=maintenance_admin_kb())
    if t == "Back to Admin Panel":
        state.pop(ADMIN_ID, None)
        return await ans(m, "Admin Panel", reply_markup=am())


def database_admin_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Download Database")],
            [kbbtn(text="Upload Database")],
            [kbbtn(text="ডিলিট অল ডাটা")],
            [kbbtn(text="Back to Admin Panel")],
        ],
        resize_keyboard=True,
    )


def _reopen_database(db_path):
    """Close the current SQLite connection and reopen the restored database."""
    global db, c
    try:
        db.commit()
    except Exception:
        pass
    try:
        db.close()
    except Exception:
        pass
    db = sqlite3.connect(db_path, check_same_thread=False)
    db.row_factory = sqlite3.Row
    c = db.cursor()

    # Re-run the base schema so a backup made from an older bot version can
    # still be opened safely. Existing tables/data are never dropped.
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users(user_id INTEGER PRIMARY KEY,username TEXT DEFAULT '',balance REAL DEFAULT 0,total_income REAL DEFAULT 0,today_income REAL DEFAULT 0,today_date TEXT DEFAULT '',created_at TEXT DEFAULT '',blocked INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,price REAL DEFAULT 0,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS submissions(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,username TEXT,category_id INTEGER,category_name TEXT,price REAL,rows_count INTEGER,estimated_profit REAL,file_name TEXT,file_path TEXT,status TEXT DEFAULT 'pending',created_at TEXT);
    CREATE TABLE IF NOT EXISTS withdrawal_categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,minimum REAL DEFAULT 0,maximum REAL DEFAULT 0,fee REAL DEFAULT 0,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS withdrawals(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,username TEXT,category_id INTEGER,category_name TEXT,amount REAL,wallet TEXT,fee REAL DEFAULT 0,total_debited REAL DEFAULT 0,status TEXT DEFAULT 'pending',created_at TEXT);
    CREATE TABLE IF NOT EXISTS support(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT,text TEXT DEFAULT '',updated_at TEXT);
    CREATE TABLE IF NOT EXISTS support_buttons(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,url TEXT NOT NULL,active INTEGER DEFAULT 1,created_at TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS force_join_channels(id INTEGER PRIMARY KEY AUTOINCREMENT,chat_id TEXT UNIQUE,channel_name TEXT,invite_link TEXT,active INTEGER DEFAULT 1,created_at TEXT);
    CREATE TABLE IF NOT EXISTS bot_settings(key TEXT PRIMARY KEY,value TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS file_reviews(submission_id INTEGER PRIMARY KEY,accepted_count INTEGER DEFAULT 0,rejected_count INTEGER DEFAULT 0,final_amount REAL DEFAULT 0,note TEXT DEFAULT '',status TEXT DEFAULT 'reviewing',updated_at TEXT DEFAULT '',admin_file_message_id INTEGER DEFAULT 0,pending_list_message_id INTEGER DEFAULT 0,review_prompt_message_id INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS custom_buttons(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE,target TEXT,active INTEGER DEFAULT 1,created_at TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS reports(id INTEGER PRIMARY KEY AUTOINCREMENT,file_name TEXT NOT NULL,file_path TEXT NOT NULL,uploaded_at REAL NOT NULL,expires_at REAL NOT NULL,green_count INTEGER DEFAULT 0,processed_users INTEGER DEFAULT 0,total_reward REAL DEFAULT 0,category_name TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS paid_uids(uid TEXT PRIMARY KEY,user_id INTEGER NOT NULL,submission_id INTEGER NOT NULL,report_id INTEGER NOT NULL,amount REAL DEFAULT 0,paid_at TEXT NOT NULL);
    """)
    migrations = [
        ("users", "blocked", "INTEGER DEFAULT 0"),
        ("users", "full_name", "TEXT DEFAULT ''"),
        ("withdrawals", "full_name", "TEXT DEFAULT ''"),
        ("withdrawals", "fee", "REAL DEFAULT 0"),
        ("withdrawals", "total_debited", "REAL DEFAULT 0"),
        ("withdrawal_categories", "fee", "REAL DEFAULT 0"),
        ("submissions", "file_hash", "TEXT DEFAULT ''"),
        ("submissions", "file_id", "TEXT DEFAULT ''"),
        ("submissions", "receiver_message_id", "INTEGER DEFAULT 0"),
        ("reports", "category_name", "TEXT DEFAULT ''"),
        ("file_reviews", "admin_file_message_id", "INTEGER DEFAULT 0"),
        ("file_reviews", "pending_list_message_id", "INTEGER DEFAULT 0"),
        ("file_reviews", "review_prompt_message_id", "INTEGER DEFAULT 0"),
    ]
    for table, col, definition in migrations:
        try:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")
        except sqlite3.OperationalError:
            pass
    db.commit()


async def database_download(m):
    if m.from_user.id != ADMIN_ID:
        return
    try:
        db.commit()
        # sqlite3.Connection.backup() creates a consistent snapshot even when
        # SQLite is using WAL/journal mode.
        with tempfile.NamedTemporaryFile(prefix="database_backup_", suffix=".db", delete=False) as tf:
            tmp_path = tf.name
        backup_db = sqlite3.connect(tmp_path)
        db.backup(backup_db)
        backup_db.commit()
        backup_db.close()
        data = Path(tmp_path).read_bytes()
        Path(tmp_path).unlink(missing_ok=True)
        filename = f"database_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        await senddoc(m.bot, ADMIN_ID, BufferedInputFile(data, filename=filename),
                      caption="✅ Database backup তৈরি হয়েছে। এই ফাইলটি নিরাপদে সংরক্ষণ করুন।")
        return await ans(m, "✅ Database সফলভাবে download করা হয়েছে।", reply_markup=database_admin_kb())
    except Exception as e:
        print(f"database download error: {e}")
        return await ans(m, "❌ Database download করতে সমস্যা হয়েছে।", reply_markup=database_admin_kb())


async def database_upload_doc(m):
    if m.from_user.id != ADMIN_ID or state.get(ADMIN_ID, {}).get("action") != "database_upload":
        return False
    fn = (m.document.file_name or "").lower()
    if not fn.endswith(".db"):
        await ans(m, "❌ শুধুমাত্র .db database file upload করুন।", reply_markup=database_admin_kb())
        return True
    tmp_path = None
    db_path = Path(DB).resolve()
    backup_path = db_path.with_name(db_path.name + ".before_restore.bak")
    try:
        f = await m.bot.get_file(m.document.file_id)
        b = io.BytesIO()
        await m.bot.download_file(f.file_path, b)
        raw = b.getvalue()

        with tempfile.NamedTemporaryFile(prefix="database_upload_", suffix=".db", delete=False) as tf:
            tmp_path = tf.name
            tf.write(raw)

        test = sqlite3.connect(tmp_path)
        test.row_factory = sqlite3.Row
        integrity = test.execute("PRAGMA integrity_check").fetchone()[0]
        tables = {r[0] for r in test.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        test.close()
        required = {"users", "submissions", "withdrawals", "categories"}
        if integrity != "ok" or not required.issubset(tables):
            raise ValueError("invalid database schema")

        db.commit()
        if db_path.exists():
            shutil.copy2(db_path, backup_path)

        # Close the active connection before atomically replacing the database file.
        try:
            db.close()
        except Exception:
            pass
        os.replace(tmp_path, db_path)
        tmp_path = None
        _reopen_database(str(db_path))
        state.pop(ADMIN_ID, None)
        return await ans(
            m,
            "✅ Database সফলভাবে upload ও restore হয়েছে।\n\nপুরোনো database-এর safety backup রাখা হয়েছে।",
            reply_markup=am()
        )
    except Exception as e:
        print(f"database upload error: {e}")
        # If replacement happened or the old connection was closed, always
        # reopen the configured database so the bot remains usable.
        try:
            _reopen_database(str(db_path))
        except Exception:
            pass
        return await ans(m, "❌ Database restore করা যায়নি। Uploaded file পরীক্ষা করুন। বর্তমান data পরিবর্তন করা হয়নি।", reply_markup=database_admin_kb())
    finally:
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass


def _clear_all_database_data():
    """Delete all rows while preserving the SQLite schema for a clean bot reset."""
    global db, c
    tables = [
        "users", "categories", "submissions", "withdrawal_categories",
        "withdrawals", "support", "support_buttons", "force_join_channels",
        "bot_settings", "file_reviews", "custom_buttons", "reports", "paid_uids"
    ]
    try:
        db.execute("BEGIN")
        for table in tables:
            db.execute(f"DELETE FROM {table}")
        # Reset AUTOINCREMENT counters where present. This is optional but makes
        # a full reset behave like a fresh database without dropping the schema.
        try:
            db.execute("DELETE FROM sqlite_sequence")
        except sqlite3.OperationalError:
            pass
        db.commit()
    except Exception:
        db.rollback()
        raise


async def database_delete_all_data(m):
    if m.from_user.id != ADMIN_ID:
        return
    action = state.get(ADMIN_ID, {}).get("action")
    if action != "database_delete_confirm":
        return
    t = clean_button_text(m.text)
    if t == "ডিলিট অল ডাটা":
        state[ADMIN_ID] = {"action": "database_delete_confirm"}
        return await ans(
            m,
            "⚠️ আপনি কি সত্যিই সব Database Data মুছে ফেলতে চান?",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[kbbtn(text="YES"), kbbtn(text="CANCEL")]],
                resize_keyboard=True,
            ),
        )
    if t == "CANCEL":
        state.pop(ADMIN_ID, None)
        return await ans(m, "❌ বাতিল করা হয়েছে।", reply_markup=database_admin_kb())
    if t == "YES":
        try:
            _clear_all_database_data()
            state.pop(ADMIN_ID, None)
            return await ans(m, "✅ সব Database Data সফলভাবে মুছে ফেলা হয়েছে।", reply_markup=database_admin_kb())
        except Exception as e:
            print(f"database delete-all error: {e}")
            state.pop(ADMIN_ID, None)
            return await ans(m, "❌ Database reset করা যায়নি।", reply_markup=database_admin_kb())
    return await ans(m, "YES অথবা CANCEL নির্বাচন করুন।", reply_markup=ReplyKeyboardMarkup(
        keyboard=[[kbbtn(text="YES"), kbbtn(text="CANCEL")]], resize_keyboard=True
    ))


async def database_admin_action(m):
    if m.from_user.id != ADMIN_ID:
        return
    t = clean_button_text(m.text)
    a = state.get(ADMIN_ID, {}).get("action")
    if t == "Database":
        state.pop(ADMIN_ID, None)
        return await ans(m, "🗄 Database Management", reply_markup=database_admin_kb())
    if t == "Back to Admin Panel":
        state.pop(ADMIN_ID, None)
        return await ans(m, "Admin Panel", reply_markup=am())
    if t == "Download Database":
        state.pop(ADMIN_ID, None)
        return await database_download(m)
    if t == "Upload Database":
        state[ADMIN_ID] = {"action": "database_upload"}
        return await ans(m, "📤 আপনার backup `.db` file পাঠান।\n\n⚠️ Restore করলে বর্তমান database-এর একটি automatic safety backup রাখা হবে।", reply_markup=database_admin_kb())
    if t == "ডিলিট অল ডাটা":
        state[ADMIN_ID] = {"action": "database_delete_confirm"}
        return await ans(
            m,
            "⚠️ আপনি কি সত্যিই সব Database Data মুছে ফেলতে চান?",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[kbbtn(text="YES"), kbbtn(text="CANCEL")]],
                resize_keyboard=True,
            ),
        )
    if a == "database_delete_confirm":
        return await database_delete_all_data(m)
    if a == "database_upload":
        return await ans(m, "📤 এখন আপনার `.db` database file পাঠান।", reply_markup=database_admin_kb())
    return await ans(m, "একটি option নির্বাচন করুন।", reply_markup=database_admin_kb())


def new_button_admin_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [kbbtn(text="Add New Button")],
            [kbbtn(text="View Added Button")],
            [kbbtn(text="Edit Added Button"),kbbtn(text="Delete Button")],
            [kbbtn(text="Back to Admin Panel")]
        ],resize_keyboard=True,
    )

async def new_button_action(m):
    if m.from_user.id!=ADMIN_ID:return
    t=clean_button_text(m.text)
    a=state.get(ADMIN_ID,{}).get("action")
    if t=="Back to Admin Panel":
        state.pop(ADMIN_ID,None)
        return await ans(m, "Admin Panel",reply_markup=am())
    if t=="New Button":
        return await ans(m, "🆕 New Button",reply_markup=new_button_admin_kb())
    if t=="Add New Button":
        state[ADMIN_ID]={"action":"add_button_name"}
        return await ans(m, "নতুন Button-এর নাম পাঠান।")
    if t=="View Added Button":
        rows=c.execute("SELECT * FROM custom_buttons WHERE active=1 ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "কোনো New Button যোগ করা হয়নি।",reply_markup=new_button_admin_kb())
        text="📋 Added Buttons\n\n"
        for r in rows:
            text += f"ID: {r['id']}\nName: {r['name']}\nTarget: {r['target']}\n\n"
        return await ans(m, text,reply_markup=new_button_admin_kb())
    if t=="Edit Added Button":
        rows=c.execute("SELECT id,name FROM custom_buttons WHERE active=1 ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "কোনো New Button নেই।",reply_markup=new_button_admin_kb())
        listing="\n".join(f"{r['id']}. {r['name']}" for r in rows)
        state[ADMIN_ID]={"action":"edit_button_id"}
        return await ans(m, "যে Button edit করবেন তার ID পাঠান।\n\n"+listing)
    if t=="Delete Button":
        rows=c.execute("SELECT id,name FROM custom_buttons WHERE active=1 ORDER BY id").fetchall()
        if not rows:
            return await ans(m, "কোনো New Button নেই।",reply_markup=new_button_admin_kb())
        listing="\n".join(f"{r['id']}. {r['name']}" for r in rows)
        state[ADMIN_ID]={"action":"delete_button_id"}
        return await ans(m, "যে Button delete করবেন তার ID পাঠান।\n\n"+listing)
    if a=="edit_button_id":
        try: bid=int(m.text.strip())
        except: return await ans(m, "সঠিক Button ID দিন।")
        r=c.execute("SELECT * FROM custom_buttons WHERE id=? AND active=1",(bid,)).fetchone()
        if not r:return await ans(m, "Button পাওয়া যায়নি। আবার ID দিন।")
        state[ADMIN_ID]={"action":"edit_button_name","id":bid,"old_name":r['name']}
        return await ans(m, f"বর্তমান নাম: {r['name']}\nনতুন Button-এর নাম পাঠান।")
    if a=="edit_button_name":
        name=m.text.strip()
        if not name or len(name)>50:return await ans(m, "সঠিক Button name দিন।")
        if c.execute("SELECT id FROM custom_buttons WHERE name=? AND id!=?",(name,state[ADMIN_ID]['id'])).fetchone():
            return await ans(m, "এই নামের Button আগে থেকেই আছে। অন্য নাম দিন।")
        state[ADMIN_ID]={"action":"edit_button_target","id":state[ADMIN_ID]['id'],"name":name}
        return await ans(m, "এখন নতুন Link অথবা User ID পাঠান।")
    if a=="edit_button_target":
        target=m.text.strip()
        if not target:return await ans(m, "Link অথবা User ID দিন।")
        bid=state[ADMIN_ID]['id']; name=state[ADMIN_ID]['name']
        c.execute("UPDATE custom_buttons SET name=?,target=?,active=1 WHERE id=?",(name,target,bid));db.commit();state.pop(ADMIN_ID,None)
        return await ans(m, f"✅ Button '{name}' update হয়েছে।",reply_markup=new_button_admin_kb())
    if a=="delete_button_id":
        try: bid=int(m.text.strip())
        except: return await ans(m, "সঠিক Button ID দিন।")
        r=c.execute("SELECT * FROM custom_buttons WHERE id=? AND active=1",(bid,)).fetchone()
        if not r:return await ans(m, "Button পাওয়া যায়নি। আবার ID দিন।")
        c.execute("UPDATE custom_buttons SET active=0 WHERE id=?",(bid,));db.commit();state.pop(ADMIN_ID,None)
        return await ans(m, f"🗑 Button '{r['name']}' delete হয়েছে।",reply_markup=new_button_admin_kb())
    return await ans(m, "একটি option নির্বাচন করুন।",reply_markup=new_button_admin_kb())


async def future_reports(m):
    if m.from_user.id!=ADMIN_ID:
        return
    reports=active_reports()
    if not reports:
        return await ans(m, "📚 Future Reports\n\nবর্তমানে কোনো active report নেই।",reply_markup=am())
    lines=["📚 Future Reports\n"]
    for i,r in enumerate(reports,1):
        left=max(0,int(float(r["expires_at"])-time.time()))
        hrs=left//3600; mins=(left%3600)//60
        lines.append(
            f"{i}. 📁 {r['file_name']}\n"
            f"   📂 Category: {r['category_name'] or 'N/A'}\n"
            f"   🟢 Green/Valid UID: {r['green_count']}\n"
            f"   👥 Processed Users: {r['processed_users']}\n"
            f"   💰 Total Reward: {float(r['total_reward'] or 0):.2f}\n"
            f"   ⏳ Expires in: {hrs}h {mins}m\n"
        )
    return await ans(m, "\n".join(lines),reply_markup=am())

async def report_admin_action(m):
    if m.from_user.id!=ADMIN_ID:
        return
    t=clean_button_text(m.text)
    a=state.get(ADMIN_ID,{}).get("action")
    if t in {"Future Reports"}:
        return await future_reports(m)
    if t=="Upload Report File":
        cleanup_expired_report()
        kb,rows=report_categories_kb()
        if not rows:
            return await ans(m, "❌ আগে Work/Category Management থেকে অন্তত একটি active category তৈরি করুন।",reply_markup=am())
        state[ADMIN_ID]={"action":"upload_report_category"}
        return await ans(m, 
            "📊 Report Upload\n\n"
            "আপনি কোন কাজ/ক্যাটাগরির রিপোর্ট আপলোড করছেন তা নির্বাচন করুন:",
            reply_markup=kb
        )
    if a=="upload_report_category":
        if t=="Cancel":
            state.pop(ADMIN_ID,None)
            return await ans(m, "Report upload cancelled.",reply_markup=am())
        rows=c.execute("SELECT * FROM categories WHERE active=1 ORDER BY id").fetchall()
        match=next((x for x in rows if x["name"]==t),None)
        if not match:
            return await ans(m, "❌ একটি valid Work/Category নির্বাচন করুন।")
        state[ADMIN_ID]={"action":"upload_report","category_name":match["name"],"category_id":match["id"]}
        return await ans(m, 
            f"✅ Category: {match['name']}\n\nএখন XLSX report file পাঠান।\nReport ১৮ ঘণ্টা active থাকবে।",
            reply_markup=ReplyKeyboardRemove()
        )
    if a=="upload_report" and not m.document:
        return await ans(m, "শুধুমাত্র XLSX report file পাঠান।")

async def report_doc(m):
    if m.from_user.id!=ADMIN_ID or state.get(ADMIN_ID,{}).get("action")!="upload_report":
        return False
    fn=m.document.file_name or "report.xlsx"
    if not fn.lower().endswith(".xlsx"):
        await ans(m, "❌ শুধুমাত্র .xlsx report গ্রহণযোগ্য।")
        return True
    try:
        f=await m.bot.get_file(m.document.file_id)
        b=io.BytesIO()
        await m.bot.download_file(f.file_path,b)
        raw=b.getvalue()
        green=extract_report_green_uids(raw)
    except Exception:
        await ans(m, "❌ Report XLSX পড়া যায়নি।")
        return True

    # Keep all active reports. Only reports older than 18 hours are removed.
    cleanup_expired_report()

    report_path=FD/f"report_{int(time.time())}_{Path(fn).name}"
    report_path.write_bytes(raw)
    uploaded=time.time(); expires=uploaded+18*60*60
    selected_category=state.get(ADMIN_ID,{}).get("category_name","")
    c.execute("INSERT INTO reports(file_name,file_path,uploaded_at,expires_at,green_count,category_name) VALUES(?,?,?,?,?,?)",
              (fn,str(report_path),uploaded,expires,len(green),selected_category))
    rid=c.lastrowid
    db.commit()
    state.pop(ADMIN_ID,None)

    results,total_reward,processed_users=await process_report(m.bot,rid,green,selected_category)

    # The report has now been processed. Remove the automatic account-upload
    # records/files so the old account queue does not remain in the bot.
    # Payment history (paid_uids) and user balances are intentionally kept.
    processed_uploads=c.execute(
        "SELECT id,file_path FROM submissions "
        "WHERE status IN ('approved','processed') AND category_name=?",
        (selected_category,)
    ).fetchall()
    for upload in processed_uploads:
        try:
            Path(upload["file_path"]).unlink(missing_ok=True)
        except Exception:
            pass
        c.execute("DELETE FROM submissions WHERE id=?",(upload["id"],))
    db.commit()

    await ans(m, 
        f"✅ Report uploaded and processed successfully.\n\n"
        f"Green/Valid UID: {len(green)}\n"
        f"Processed Users: {processed_users}\n"
        f"Total Reward Added: {total_reward:.2f}\n"
        f"Valid for: 18 hours",
        reply_markup=am()
    )

    # Notify all registered users that a new report is available. The file itself
    # is intentionally NOT pushed here; users can open the Report button to receive it.
    category_text=selected_category or "Report"
    notification=(
        "📊 রিপোর্ট এসে গেছে!\n\n"
        f"📁 Category: {category_text}\n"
        "📊 Report বাটনে ক্লিক করে আপনার রিপোর্ট চেক করতে পারেন।"
    )
    all_users=c.execute("SELECT user_id FROM users WHERE blocked=0 ORDER BY user_id").fetchall()
    sent=0
    for row in all_users:
        uid=row["user_id"]
        try:
            await sendmsg(m.bot, uid,notification,reply_markup=um(uid))
            sent+=1
        except Exception:
            pass
    await ans(m, f"📢 Report notification sent to {sent} users.",reply_markup=am())
    return True

async def admin_config_action(m):
    if m.from_user.id != ADMIN_ID:
        return
    a=state.get(ADMIN_ID,{}).get("action")
    if a not in {"add_button_name","add_button_target","receiver_channel"}:
        return
    text=(m.text or "").strip()
    try:
        if a=="add_button_name":
            if not text or len(text)>50: raise ValueError()
            state[ADMIN_ID]={"action":"add_button_target","name":text}
            return await ans(m, "এখন Button-এর Link অথবা User ID পাঠান।\nউদাহরণ: https://t.me/example অথবা 123456789")
        if a=="add_button_target":
            name=state[ADMIN_ID]["name"]
            if not text: raise ValueError()
            c.execute("INSERT INTO custom_buttons(name,target,active,created_at) VALUES(?,?,1,?) ON CONFLICT(name) DO UPDATE SET target=excluded.target,active=1",(name,text,now()))
            db.commit(); state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Button '{name}' সফলভাবে যোগ/আপডেট হয়েছে।",reply_markup=am())
        if a=="receiver_channel":
            if text.upper()=="OFF":
                save_setting("receiver_channel",""); state.pop(ADMIN_ID,None)
                return await ans(m, "✅ Receiver Channel বন্ধ করা হয়েছে। User file Admin Panel-এ যাবে না; channel না থাকলে upload গ্রহণ করা হবে না।",reply_markup=am())
            target=normalize_receiver_target(text)
            if not target: raise ValueError()
            try:
                await m.bot.get_chat(target)
            except Exception:
                return await ans(m, "❌ Channel পাওয়া যাচ্ছে না। Public @username বা https://t.me/username দিন এবং bot-কে channel-এ admin করুন।")
            save_setting("receiver_channel",target); state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Receiver Channel saved: {target}\n\nএখন user-এর uploaded file সরাসরি এই channel-এ যাবে।",reply_markup=am())
    except Exception:
        return await ans(m, "❌ তথ্যটি সঠিক নয়। আবার চেষ্টা করুন।")

async def admin(m):
    if m.from_user.id!=ADMIN_ID:return
    t=clean_button_text(m.text); s=state.get(ADMIN_ID,{})

    # Main navigation must cancel any stale withdrawal state first.
    if t in ("Admin Panel", "User Menu"):
        state.pop(ADMIN_ID, None)
        if t == "Admin Panel":
            return await ans(m, "Admin Panel", reply_markup=am())
        return await ans(m, "User Menu", reply_markup=um(ADMIN_ID))

    # Custom User Panel buttons must work for the admin too when testing the User Menu.
    # A ReplyKeyboard button cannot directly open a URL, so convert the click into an inline URL button.
    if await handle_custom_button(m):
        return

    # User-panel multi-step inputs must be handled before admin menu routing.
    # In particular, withdrawal amount/wallet messages must not be swallowed
    # by the admin category/menu handler when the admin tests the User Panel.
    if s.get("withdraw_step") == "amount" or s.get("withdraw_step") == "wallet":
        return await user_state(m)

    # Handle category selection only when the state is actually waiting for
    # a category button.
    if s.get("cat"):
        return await sell_category_selection(m)
    if s.get("wcat") and not s.get("withdraw_step"):
        return await withdrawal_category_selection(m)

    # Admin can also test the complete User Panel. These buttons must not be
    # swallowed by the generic admin handler.
    if t=="Account":
        return await account(m)
    if t in ("Report","Pending Account","📊 Report"):
        return await report_section(m)
    if t=="Sell Account":
        return await sell(m)
    if t=="Withdrawal":
        return await withdraw(m)
    if t=="Price List":
        return await prices(m)
    if t=="Support":
        return await support(m)

    if t=="User Management":
        state.pop(ADMIN_ID,None)
        return await user_management(m)
    if t in ("Add Balance","Remove Balance"):
        state[ADMIN_ID]={"action":"add" if t=="Add Balance" else "remove"}
        return await ans(m, "Format: USER_ID AMOUNT", reply_markup=user_mgmt())
    if t=="Work Management":
        state[ADMIN_ID]={"action":"work_menu"}
        return await ans(m, 
            "Work / Price Management",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [kbbtn(text="Add Work"),kbbtn(text="Edit Work")],
                    [kbbtn(text="Disable Work"),kbbtn(text="Enable Work")],
                    [kbbtn(text="Disable Notice")],
                    [kbbtn(text="Work List")],
                    [kbbtn(text="Back to Admin Panel")]
                ],resize_keyboard=True,
            )
        )
    if t=="Withdrawal Management":
        state.pop(ADMIN_ID,None)
        return await ans(m, 
            "💳 Withdrawal Management\n\n"
            "এখান থেকে Withdrawal category, limit এবং pending request manage করুন।",
            reply_markup=withdrawal_admin_kb()
        )
    if t=="Pending Files":
        rows=c.execute("SELECT * FROM submissions WHERE status='pending' ORDER BY id DESC").fetchall()
        if not rows:
            return await ans(m, "📂 Pending Files\n\nকোনো pending file নেই।",reply_markup=am())
        sent=0
        unavailable=0
        for s in rows:
            existing=c.execute("SELECT pending_list_message_id FROM file_reviews WHERE submission_id=?",(s["id"],)).fetchone()
            if existing and existing["pending_list_message_id"]:
                continue

            caption=(f"New Account File\n\nSubmission ID: {s['id']}\n"
                     f"Username: @{s['username'] or 'N/A'}\nUser ID: {s['user_id']}\n"
                     f"Work: {s['category_name']}\nAccounts: {s['rows_count']}\n"
                     f"Price: {s['price']:.2f}\nEstimated Profit: {s['estimated_profit']:.2f}\n"
                     f"File: {s['file_name']}\nTime: {s['created_at']}")
            kb=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🔍 Review File",callback_data=f"review:{s['id']}"),
                InlineKeyboardButton(text="❌ Reject File",callback_data=f"reject_file:{s['id']}")
            ]])

            delivered=False
            # 1) Prefer Telegram file_id: survives Railway restarts/deploys.
            try:
                fid = s["file_id"]
            except Exception:
                fid = None
            if fid:
                try:
                    msg=await m.bot.send_document(
                        ADMIN_ID, fid, caption=caption, reply_markup=kb
                    )
                    delivered=True
                except Exception as e:
                    print(f"Pending file #{s['id']} file_id send error: {e}")

            # 2) Fall back to the local file if it still exists.
            if not delivered:
                try:
                    p=Path(s["file_path"])
                    if p.exists():
                        msg=await senddoc(
                            m.bot, ADMIN_ID,
                            BufferedInputFile(p.read_bytes(),filename=s["file_name"]),
                            caption=caption, reply_markup=kb
                        )
                        delivered=True
                except Exception as e:
                    print(f"Pending file #{s['id']} local send error: {e}")

            # 3) Final recovery: copy the original receiver-channel message.
            if not delivered:
                try:
                    rid = int(s["receiver_message_id"] or 0)
                except Exception:
                    rid = 0
            if not delivered and rid:
                receiver=setting("receiver_channel").strip()
                if receiver:
                    try:
                        msg=await m.bot.copy_message(
                            chat_id=ADMIN_ID,
                            from_chat_id=receiver,
                            message_id=rid,
                            caption=caption,
                            reply_markup=kb
                        )
                        delivered=True
                    except Exception as e:
                        print(f"Pending file #{s['id']} receiver copy error: {e}")

            if delivered:
                try:
                    c.execute(
                        "INSERT INTO file_reviews(submission_id,status,updated_at,pending_list_message_id) "
                        "VALUES(?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET "
                        "pending_list_message_id=excluded.pending_list_message_id",
                        (s["id"],"pending",now(),msg.message_id)
                    )
                    db.commit()
                except Exception as e:
                    print(f"Pending file #{s['id']} review tracking error: {e}")
                sent += 1
            else:
                unavailable += 1
                # Never silently hide a pending submission. Show its metadata and
                # leave the record pending so the admin can still see/reject it.
                try:
                    msg=await ans(
                        m,
                        caption + "\n\n⚠️ File could not be recovered from Telegram or local storage.",
                        reply_markup=kb
                    )
                    c.execute(
                        "INSERT INTO file_reviews(submission_id,status,updated_at,pending_list_message_id) "
                        "VALUES(?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET "
                        "pending_list_message_id=excluded.pending_list_message_id",
                        (s["id"],"pending",now(),msg.message_id)
                    )
                    db.commit()
                except Exception as e:
                    print(f"Pending file #{s['id']} metadata send error: {e}")
        return await ans(
            m,
            f"📂 Pending Files\n\nShowing {sent} pending file(s)."
            + (f"\nUnavailable: {unavailable}" if unavailable else ""),
            reply_markup=am()
        )
    if t=="Pending Withdrawals":
        r=c.execute("SELECT * FROM withdrawals WHERE status='pending' ORDER BY id DESC").fetchall()
        if not r:
            return await ans(m, "Pending Withdrawals: 0\n\nকোনো pending withdrawal নেই।",reply_markup=am())
        for x in r:
            text=(f"💳 Withdrawal #{x['id']}\n\n"
                  f"Account Name: N/A\n"
                  f"Username: @{x['username'] or 'N/A'}\n"
                  f"User ID: {x['user_id']}\n"
                  f"Category: {x['category_name']}\n"
                  f"Amount: {x['amount']:.2f}\n"
                  f"Fee: {float(x['fee'] or 0):.2f} টাকা\n"
                  f"Total Debited: {float(x['total_debited'] or x['amount']):.2f} টাকা\n"
                  f"Payment Number: {x['wallet']}\n"
                  f"Time: {x['created_at']} UTC")
            kb=InlineKeyboardMarkup(inline_keyboard=[[
                ibtn("Approve", callback_data=f"wapprove:{x['id']}", icon_custom_emoji_id=PE["done"]),
                ibtn(text="❌ Reject",callback_data=f"wreject:{x['id']}")
            ]])
            await ans(m, text,reply_markup=kb)
        return await ans(m, f"Pending Withdrawals: {len(r)}",reply_markup=am())
    if t=="Set Payment Channel":
        current=setting("payment_approve_channel")
        state[ADMIN_ID]={"action":"payment_approve_channel"}
        return await ans(m, "Set Payment Channel-এর @username, public channel link অথবা numeric Channel ID পাঠান.\nউদাহরণ: @mychannel অথবা https://t.me/mychannel অথবা -1001234567890\n\nবন্ধ করতে OFF লিখুন.\nবর্তমান: "+(current or "Not set"))
    if t=="Generate Main Excel":return await merge(m)
    if t=="New Button":
        state.pop(ADMIN_ID,None)
        return await ans(m, "🆕 New Button",reply_markup=new_button_admin_kb())
    if t=="Set Receiver Channel":
        current=setting("receiver_channel")
        state[ADMIN_ID]={"action":"receiver_channel"}
        return await ans(m, "Receiver Channel-এর @username বা channel link পাঠান.\nউদাহরণ: @mychannel অথবা https://t.me/mychannel\n\n" f"বর্তমান: {current or 'Not set'}\n\nবন্ধ করতে OFF লিখুন.")
    if t=="Broadcast":state[ADMIN_ID]={"action":"broadcast"};return await ans(m, "Broadcast message পাঠান.")
    if t=="Support Management":
        state.pop(ADMIN_ID,None)
        count=c.execute("SELECT COUNT(*) n FROM support_buttons WHERE active=1").fetchone()["n"]
        return await ans(
            m,
            "📞 Support Management\n\n"
            f"Support List: {count}টি active button\n\n"
            "নিচের ৩টি অপশন থেকে নির্বাচন করুন।",
            reply_markup=support_admin_kb()
        )

    a=s.get("action")
    try:
        if a=="fj_chat_id":
            state[ADMIN_ID]={"action":"fj_name","chat_id":m.text.strip()}
            return await ans(m, "ধাপ ২/৩\nChannel-এর নাম পাঠান।",reply_markup=force_join_admin_kb())
        if a=="fj_name":
            s=state[ADMIN_ID];s["channel_name"]=m.text.strip();s["action"]="fj_link";state[ADMIN_ID]=s
            return await ans(m, "ধাপ ৩/৩\nChannel-এর Join/Invite Link পাঠান।",reply_markup=force_join_admin_kb())
        if a=="fj_link":
            s=state[ADMIN_ID]
            c.execute("INSERT INTO force_join_channels(chat_id,channel_name,invite_link,active,created_at) VALUES(?,?,?,?,?)",
                      (s["chat_id"],s["channel_name"],m.text.strip(),1,now()))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Force Join channel সফলভাবে যোগ হয়েছে।",reply_markup=force_join_admin_kb())
        if a=="fj_remove":
            cid=int(m.text.strip())
            c.execute("UPDATE force_join_channels SET active=0 WHERE id=?",(cid,))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Force Join channel remove হয়েছে।",reply_markup=force_join_admin_kb())

        if a=="disable_notice":
            notice_text = m.text if m.text is not None else m.caption
            notice_entities = m.entities if m.text is not None else m.caption_entities
            if not notice_text:
                return await ans(m, "❌ Disable Notice হিসেবে একটি text message পাঠান।")
            save_setting("disable_notice_text", notice_text)
            save_setting("disable_notice_entities", json.dumps(_serialize_entities(notice_entities), ensure_ascii=False))
            state.pop(ADMIN_ID,None)
            return await ans(m, "✅ Disable Notice সফলভাবে save হয়েছে। Premium/Custom Emoji-সহ message সংরক্ষণ করা হয়েছে।", reply_markup=am())

        if a=="work_name":
            state[ADMIN_ID]={"action":"work_price","name":m.text.strip()}
            return await ans(m, "এই Work-এর প্রতি Account Price পাঠান।")
        if a=="work_price":
            s=state[ADMIN_ID]
            c.execute("INSERT INTO categories(name,price,active) VALUES(?,?,1)",(s["name"],float(m.text.strip())))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Work সফলভাবে যোগ হয়েছে।",reply_markup=am())
        if a=="work_edit_id":
            cid=int(m.text.strip())
            state[ADMIN_ID]={"action":"work_edit_name","id":cid}
            return await ans(m, "নতুন Work Name পাঠান।")
        if a=="work_edit_name":
            s=state[ADMIN_ID];s["name"]=m.text.strip();s["action"]="work_edit_price";state[ADMIN_ID]=s
            return await ans(m, "নতুন Price পাঠান।")
        if a=="work_edit_price":
            s=state[ADMIN_ID]
            c.execute("UPDATE categories SET name=?,price=? WHERE id=?",(s["name"],float(m.text.strip()),s["id"]))
            db.commit();state.pop(ADMIN_ID,None)
            return await ans(m, "Work update হয়েছে।",reply_markup=am())
        if a=="work_toggle_id":
            cid=int(m.text.strip());toggle=state[ADMIN_ID]["toggle"]
            cat=c.execute("SELECT * FROM categories WHERE id=?",(cid,)).fetchone()
            if not cat:
                raise ValueError()
            was_active=int(cat["active"] or 0) == 1
            c.execute("UPDATE categories SET active=? WHERE id=?",(toggle,cid))
            db.commit();state.pop(ADMIN_ID,None)
            if toggle==0 and was_active:
                sent,failed=await notify_work_disabled(m.bot, cat["name"])
                return await ans(m,
                    f"🔕 {cat['name']} Work বন্ধ করা হয়েছে।\n\n"
                    f"Notification Sent: {sent} | Failed: {failed}",
                    reply_markup=am())
            return await ans(m, "Work status update হয়েছে।",reply_markup=am())

        if a=="review_accepted":
            sid=state[ADMIN_ID]["sid"]
            s=c.execute("SELECT * FROM submissions WHERE id=? AND status='pending'",(sid,)).fetchone()
            accepted=int(m.text.strip())
            if not s or accepted<0 or accepted>s["rows_count"]: raise ValueError()
            rejected=s["rows_count"]-accepted
            amount=accepted*float(s["price"])
            c.execute("INSERT INTO file_reviews(submission_id,accepted_count,rejected_count,final_amount,status,updated_at) VALUES(?,?,?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET accepted_count=excluded.accepted_count,rejected_count=excluded.rejected_count,final_amount=excluded.final_amount,status='reviewing',updated_at=excluded.updated_at",
                      (sid,accepted,rejected,amount,"reviewing",now()))
            db.commit()
            kb=InlineKeyboardMarkup(inline_keyboard=[
                [ibtn(text="Confirm & Add Balance",callback_data=f"review_confirm:{sid}")],
                [ibtn(text="Cancel Review",callback_data=f"review_cancel:{sid}")]
            ])
            return await ans(m, 
                f"Review Report — #{sid}\n\nSubmitted: {s['rows_count']}\nAccepted: {accepted}\nRejected: {rejected}\n"
                f"Price: {s['price']:.2f}/account\nFinal Amount: {amount:.2f}",
                reply_markup=kb
            )

        if a=="reject_reason":
            sid=state[ADMIN_ID]["sid"]
            s=c.execute("SELECT * FROM submissions WHERE id=? AND status='pending'",(sid,)).fetchone()
            if not s: raise ValueError()
            reason=m.text.strip()
            c.execute("UPDATE submissions SET status='rejected' WHERE id=?",(sid,))
            c.execute("INSERT INTO file_reviews(submission_id,rejected_count,note,status,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET rejected_count=excluded.rejected_count,note=excluded.note,status='rejected',updated_at=excluded.updated_at",
                      (sid,s["rows_count"],reason,"rejected",now()))
            r=c.execute("SELECT pending_list_message_id,admin_file_message_id FROM file_reviews WHERE submission_id=?",(sid,)).fetchone()
            db.commit();state.pop(ADMIN_ID,None)
            for mid in ((r["pending_list_message_id"] if r else 0),(r["admin_file_message_id"] if r else 0)):
                try:
                    if mid: await m.bot.delete_message(ADMIN_ID,int(mid))
                except Exception: pass
            await sendmsg(
            q.message.bot,
            s.get("user_id"),
            premiumize(
                f"Submission #{submission_num} approved.\\n\\n"
                f"5420592870396557986 Accepted: {r['accepted_count']}\\n"
                f"5778318896889074623 Rejected: {r['rejected_count']}\\n"
                f"6240150562654917631 Added Balance: {r['final_amount']:.2f}"
            ),
            reply_markup=um(s.get("user_id"))
        )
            try:
                await sendmsg(m.bot, s["user_id"],f"File Review Complete\n\nStatus: Rejected\nReason: {reason}",reply_markup=um(s["user_id"]))
            except: pass
            await cleanup_review_submission(m.bot, sid, r, s)
            return

        if a in ("add","remove"):
            uid,amt=m.text.split()[:2];uid=int(uid);amt=float(amt)
            exists=c.execute("SELECT user_id FROM users WHERE user_id=?",(uid,)).fetchone()
            if not exists: raise ValueError()
            delta=amt if a=="add" else -amt
            c.execute("UPDATE users SET balance=balance+? WHERE user_id=?",(delta,uid));db.commit()
            await ans(m, f"Balance updated: {delta:.2f}",reply_markup=user_mgmt());state.pop(ADMIN_ID,None);return
        if a=="block":
            uid=int(m.text.strip())
            exists=c.execute("SELECT user_id FROM users WHERE user_id=?",(uid,)).fetchone()
            if not exists: raise ValueError()
            c.execute("UPDATE users SET blocked=1 WHERE user_id=?",(uid,));db.commit()
            state.pop(ADMIN_ID,None)
            await ans(m, f"User {uid} blocked.",reply_markup=user_mgmt())
            try: await sendmsg(m.bot, uid,"Your account has been blocked by admin.")
            except: pass
            return
        p=m.text.split("|");cmd=p[0].upper()
        if a=="cat":
            if cmd=="ADD":c.execute("INSERT INTO categories(name,price) VALUES(?,?)",(p[1].strip(),float(p[2])))
            elif cmd=="UPDATE":c.execute("UPDATE categories SET name=?,price=? WHERE id=?",(p[2].strip(),float(p[3]),int(p[1])))
            elif cmd in ("ON","OFF"):c.execute("UPDATE categories SET active=? WHERE id=?",(1 if cmd=="ON" else 0,int(p[1])))
            else:raise ValueError()
            db.commit();await ans(m, "Work/Price updated.",reply_markup=am());state.pop(ADMIN_ID,None);return
        if a=="wc":
            if cmd=="ADD":c.execute("INSERT INTO withdrawal_categories(name,minimum,maximum,fee) VALUES(?,?,?,?)",(p[1].strip(),float(p[2]),float(p[3]),float(p[4]) if len(p)>4 else 0))
            elif cmd=="UPDATE":c.execute("UPDATE withdrawal_categories SET name=?,minimum=?,maximum=?,fee=? WHERE id=?",(p[2].strip(),float(p[3]),float(p[4]),float(p[5]) if len(p)>5 else 0,int(p[1])))
            elif cmd in ("ON","OFF"):c.execute("UPDATE withdrawal_categories SET active=? WHERE id=?",(1 if cmd=="ON" else 0,int(p[1])))
            else:raise ValueError()
            db.commit();await ans(m, "Withdrawal category updated.",reply_markup=am());state.pop(ADMIN_ID,None);return
        if a=="support_link":
            link=m.text.strip()
            if not link: raise ValueError()
            save_setting("support_link",link)
            state.pop(ADMIN_ID,None)
            return await ans(m, f"Support link saved: {link}",reply_markup=am())
        if a=="payment_approve_channel":
            target=m.text.strip()
            if target.upper()=="OFF":
                save_setting("payment_approve_channel",""); state.pop(ADMIN_ID,None)
                return await ans(m, "✅ Set Payment Channel বন্ধ করা হয়েছে।",reply_markup=am())
            target=normalize_receiver_target(target)
            if not target: raise ValueError()
            try:
                await m.bot.get_chat(target)
            except Exception:
                return await ans(m, "❌ Channel পাওয়া যায়নি। Public @username/link বা numeric Channel ID দিন এবং bot-কে channel admin করুন।")
            save_setting("payment_approve_channel",target); state.pop(ADMIN_ID,None)
            return await ans(m, f"✅ Set Payment Channel saved: {target}",reply_markup=am())
        if a=="broadcast":
            n=0
            failed=0
            # Copy the original Telegram message instead of rebuilding it from
            # m.text. This preserves Custom Emoji entities, formatting, links,
            # and other message entities exactly as the admin sent them.
            for x in c.execute("SELECT user_id FROM users").fetchall():
                try:
                    await m.bot.copy_message(
                        chat_id=x["user_id"],
                        from_chat_id=m.chat.id,
                        message_id=m.message_id,
                        reply_markup=um(x["user_id"])
                    )
                    n += 1
                except Exception:
                    failed += 1
            state.pop(ADMIN_ID,None)
            return await ans(m, f"Broadcast complete. Sent: {n} | Failed: {failed}", reply_markup=am())
    except: await ans(m, "Invalid format. আবার চেষ্টা করুন.")


def clean_button_text(text):
    text = re.sub(r"^[\u26AA\U0001F7E2\U0001F535\U0001F534\U0001F7E1\U0001F7E3]\s*", "", text or "")
    return " ".join((text or "").replace("–","—").replace("−","-").split()).strip()

async def sell_category_selection(m):
    if m.from_user.id != ADMIN_ID and not await ensure_joined(m):
        return
    s = state.get(m.from_user.id, {})
    if not s.get("cat"):
        return
    text = clean_button_text(m.text)
    if text == "Cancel":
        state.pop(m.from_user.id, None)
        return await ans(m, "Cancelled.", reply_markup=um(m.from_user.id))

    rows = c.execute("SELECT * FROM categories WHERE active=1 ORDER BY id").fetchall()
    for x in rows:
        expected = clean_button_text(f"{x['name']} — {x['price']:.2f}")
        if text == expected:
            state[m.from_user.id] = {"category_id": x["id"]}
            return await ans(m, 
                f"Work selected: {x['name']}\n"
                f"Price: {x['price']:.2f}/account\n\n"
                f"এখন XLSX file পাঠান।\n"
                f"⚠️ ফাইলের account password-এর শেষে অবশ্যই তারিখ থাকতে হবে।\n"
                f"উদাহরণ: Sayed@12 বা Rafiya26-এর মতো date suffix ব্যবহার করুন।",
                reply_markup=ReplyKeyboardRemove()
            )

    return await ans(m, "দয়া করে নিচের Work button থেকে একটি নির্বাচন করুন।")

async def withdrawal_category_selection(m):
    if m.from_user.id != ADMIN_ID and not await ensure_joined(m):
        return
    s = state.get(m.from_user.id, {})
    if not s.get("wcat"):
        return
    text = clean_button_text(m.text)
    if text == "Cancel":
        state.pop(m.from_user.id, None)
        return await ans(m, "Cancelled.", reply_markup=um(m.from_user.id))

    rows = c.execute("SELECT * FROM withdrawal_categories WHERE active=1 ORDER BY id").fetchall()
    for x in rows:
        # User-facing button contains only the category name.
        expected = clean_button_text(x["name"])
        if text == expected:
            nml=x["name"].strip().lower()
            if "বিকাশ" in nml or "bkash" in nml:
                wallet_label="বিকাশ"
            elif "নগদ" in nml or "nagad" in nml:
                wallet_label="নগদ"
            else:
                wallet_label=x["name"].strip()
            state[m.from_user.id] = {
                "wid": x["id"], "wn": x["name"],
                "min": float(x["minimum"]), "max": float(x["maximum"]),
                "fee": float(x["fee"] or 0),
                "wallet_label": wallet_label,
                "withdraw_step": "amount"
            }
            return await ans(m, 
                f"Category: {x['name']}\n"
                f"Minimum: {x['minimum']:.2f}\n"
                f"Maximum: {x['maximum']:.2f}\n"
                f"Withdrawal Fee: {float(x['fee'] or 0):.2f} টাকা\n\n"
                f"কত টাকা উত্তোলন করবেন? শুধু amount লিখুন."
            )

    return await ans(m, "দয়া করে নিচের Withdrawal category button থেকে একটি নির্বাচন করুন।")

async def user_state(m):
    if not await ensure_joined(m): return
    s=state.get(m.from_user.id,{})
    text=clean_button_text(m.text)

    # Cancel must always exit any active withdrawal/workflow state.
    if text == "Cancel":
        state.pop(m.from_user.id, None)
        return await ans(m, "Cancelled.", reply_markup=um(m.from_user.id))

    # Withdrawal amount/wallet steps must be handled BEFORE the
    # category-selection state. The category flag remains set while the
    # user enters the amount, so checking wcat first swallowed the amount.
    if s.get("withdraw_step") == "amount":
        try:
            a=float(m.text.strip())
        except (TypeError, ValueError):
            return await ans(m, "❌ সঠিক amount লিখুন। উদাহরণ: 100")
        if a <= 0:
            return await ans(m, "❌ Amount 0-এর বেশি হতে হবে।")
        u=c.execute("SELECT balance FROM users WHERE user_id=?",(m.from_user.id,)).fetchone()
        if u is None:
            user(m.from_user)
            u=c.execute("SELECT balance FROM users WHERE user_id=?",(m.from_user.id,)).fetchone()
        if a < s["min"]:
            return await ans(m, f"❌ Minimum withdrawal: {s['min']:.2f}")
        if s["max"] > 0 and a > s["max"]:
            return await ans(m, f"❌ Maximum withdrawal: {s['max']:.2f}")
        fee=float(s.get("fee",0) or 0)
        total_debited=a+fee
        if total_debited > float(u["balance"] or 0):
            return await ans(m, f"❌ Balance যথেষ্ট নয়।\nWithdrawal: {a:.2f} টাকা\nFee: {fee:.2f} টাকা\nমোট প্রয়োজন: {total_debited:.2f} টাকা\nআপনার Balance: {float(u['balance'] or 0):.2f}")
        s["withdraw_step"]="wallet"
        s["value"]=a
        s["fee"]=fee
        s["total_debited"]=total_debited
        state[m.from_user.id]=s
        return await ans(m, f"📱 আপনার {s.get('wallet_label', s.get('wn', 'wallet'))} নাম্বার দিন।")

    if s.get("withdraw_step") == "wallet":
        wallet=m.text.strip()
        if not wallet:
            return await ans(m, "❌ Wallet/number দিন।")
        # Re-check balance immediately before deduction to prevent stale state issues.
        u=c.execute("SELECT balance FROM users WHERE user_id=?",(m.from_user.id,)).fetchone()
        if not u or float(u["balance"] or 0) < float(s.get("total_debited", s["value"])):
            state.pop(m.from_user.id,None)
            return await ans(m, "❌ আপনার balance আর পর্যাপ্ত নেই। Withdrawal বাতিল করা হয়েছে।",reply_markup=um(m.from_user.id))
        c.execute("UPDATE users SET balance=balance-? WHERE user_id=?",(s.get("total_debited",s["value"]),m.from_user.id))
        full_name = " ".join(x for x in [m.from_user.first_name or "", m.from_user.last_name or ""] if x).strip()
        c.execute("INSERT INTO withdrawals(user_id,username,full_name,category_id,category_name,amount,wallet,fee,total_debited,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(m.from_user.id,m.from_user.username or "",full_name,s["wid"],s["wn"],s["value"],wallet,s.get("fee",0),s.get("total_debited",s["value"]),"pending",now()))
        wid=c.lastrowid
        db.commit()
        state.pop(m.from_user.id,None)
        await ans(m, f"✅ Withdrawal request জমা হয়েছে.\nRequest ID: {wid}\nAmount: {s['value']:.2f} টাকা\n\nAdmin approval-এর অপেক্ষায়.",reply_markup=um(m.from_user.id))
        return await sendmsg(m.bot, ADMIN_ID,f"New Withdrawal Request\n\nID: {wid}\nUsername: @{m.from_user.username or 'N/A'}\nUser ID: {m.from_user.id}\nCategory: {s['wn']}\nAmount: {s['value']:.2f} টাকা\nWallet: {wallet}\nTime: {now()}")

    # Only after the withdrawal steps are finished should a new
    # category/work selection be processed.
    if s.get("wcat"):
        return await withdrawal_category_selection(m)
    if s.get("cat"):
        return await sell_category_selection(m)

    if await handle_custom_button(m):
        return

    await ans(m, "Menu থেকে একটি অপশন নির্বাচন করুন।",reply_markup=um(m.from_user.id))



def mask_payment_number(value):
    v=str(value or "").strip()
    if len(v) <= 7:
        return v
    return v[:3] + "..." + v[-4:]

async def send_payment_approval_report(bot, w):
    target=setting("payment_approve_channel").strip()
    if not target:
        return
    try:
        total=c.execute("SELECT COALESCE(SUM(amount),0) total FROM withdrawals WHERE user_id=? AND status='approved'",(w["user_id"],)).fetchone()["total"]
        text=("✅ Payment Approved\n\n"
              f"Account Name: {w['full_name'] or 'N/A'}\n"
              f"Username: @{w['username'] or 'N/A'}\n"
              f"Withdrawal Amount: {float(w['amount']):.2f} টাকা\n"
              f"Withdrawal Fee: {float(w['fee'] or 0):.2f} টাকা\n"
              f"Total Debited: {float(w['total_debited'] or w['amount']):.2f} টাকা\n"
              f"Payment Number: {mask_payment_number(w['wallet'])}\n"
              f"Total Withdrawal: {float(total):.2f}")
        await sendmsg(bot, target,text)
    except Exception:
        pass

async def withdrawal_add_method_callback(q):
    if q.from_user.id!=ADMIN_ID:
        return await q.answer("Not allowed.", show_alert=True)
    method=q.data.split(":",1)[1]
    name="বিকাশ" if method=="bkash" else "নগদ"
    state[ADMIN_ID]={"action":"w_add_min","name":name}
    await q.answer()
    return await ans(q.message, f"Payment method: {name}\n\nMinimum withdrawal amount পাঠান।")

async def withdrawal_edit_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    wid=int(q.data.split(":")[1])
    x=c.execute("SELECT * FROM withdrawal_categories WHERE id=?",(wid,)).fetchone()
    if not x:return await q.answer("Category not found.",show_alert=True)
    state[ADMIN_ID]={"action":"w_edit_name","id":wid}
    await q.answer()
    await ans(q.message, f"Editing #{wid}\nMethod: {x['name']}\nMinimum: {float(x['minimum']):.2f}\nMaximum: {float(x['maximum']):.2f}\nFee: {float(x['fee'] or 0):.2f} টাকা\n\nনতুন Category name পাঠান।\n(বিকাশ/নগদ লিখতে পারেন।)" )

async def withdrawal_toggle_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    wid=int(q.data.split(":")[1])
    x=c.execute("SELECT active FROM withdrawal_categories WHERE id=?",(wid,)).fetchone()
    if not x:return await q.answer("Category not found.",show_alert=True)
    new=0 if x["active"] else 1
    c.execute("UPDATE withdrawal_categories SET active=? WHERE id=?",(new,wid));db.commit()
    await q.answer("Updated.")
    await q.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [ibtn(text="Edit",callback_data=f"wedit:{wid}"),
         ibtn(text="Disable" if new else "Enable",callback_data=f"wtoggle:{wid}")]
    ]))

async def withdrawal_delete_callback(q):
    if q.from_user.id!=ADMIN_ID:
        return await q.answer("Not allowed.", show_alert=True)
    try:
        wid=int(q.data.split(":")[1])
    except Exception:
        return await q.answer("Invalid category.", show_alert=True)

    x=c.execute("SELECT * FROM withdrawal_categories WHERE id=?", (wid,)).fetchone()
    if not x:
        return await q.answer("Category not found.", show_alert=True)

    # Keep historical withdrawal records intact; only remove the configured
    # payment category so it can no longer be selected for new withdrawals.
    c.execute("DELETE FROM withdrawal_categories WHERE id=?", (wid,))
    db.commit()

    await q.answer("Payment category deleted.")
    try:
        await q.message.edit_text(
            f"✅ Payment Category Deleted\\n\\n"
            f"Category: {x['name']}\\n"
            f"ID: {wid}\\n\\n"
            "এই category এখন আর নতুন withdrawal-এর জন্য available নয়।"
        )
    except Exception:
        pass

async def withdrawal_approve_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    wid=int(q.data.split(":")[1])
    w=c.execute("SELECT * FROM withdrawals WHERE id=? AND status='pending'",(wid,)).fetchone()
    if not w:return await q.answer("Already processed.",show_alert=True)
    c.execute("UPDATE withdrawals SET status='approved' WHERE id=?",(wid,));db.commit()
    await q.answer("Approved.")
    await q.message.edit_reply_markup(reply_markup=None)
    try:
        method = "Bkash" if "bkash" in str(w["category_name"] or "").lower() else "Nogod"
        success_text = (
            "{{PE:6190336264940559752}} WITHDRAWAL SUCCESSFUL\n"
            "━━━━━━━━━━━━━━━━━━━\n"
            f"{{{{PE:6066456029300788292}}}} Amount: {float(w['amount']):.2f} BDT\n"
            f"Withdrawal Fee: {float(w['fee'] or 0):.2f} BDT\n"
            f"{{{{PE:6237975191784266396}}}} Method: {method}\n"
            f"{{{{PE:5291913773307667021}}}} Request ID: #{wid}\n"
            "━━━━━━━━━━━━━━━━━━━\n\n"
            "{{PE:5951982867255924070}} Withdrawal Completed\n"
            "Your payment has been successfully sent.\n"
            "Thank you for using our service. {{PE:5474525960143385880}}\n"
            "━━━━━━━━━━━━━━━━━━━\n"
            "{{PE:5472250091332993630}} Payment source: @earntotalfresh_bot"
        )
        # Send the approval notification to the user first, then FORWARD
        # that exact Telegram message to the configured Set Payment Channel.
        # Forwarding the original message preserves its MessageEntity objects,
        # including Telegram Premium Custom Emoji, instead of rebuilding the
        # message and risking broken/missing custom emojis.
        sent = await sendmsg(q.message.bot,w["user_id"],success_text,reply_markup=um(w["user_id"]))
        target = setting("payment_approve_channel").strip()
        if target:
            try:
                await q.message.bot.forward_message(
                    chat_id=target,
                    from_chat_id=w["user_id"],
                    message_id=sent.message_id
                )
            except Exception as forward_error:
                print(f"[payment-approve-forward] withdrawal #{wid}: {forward_error}")
    except Exception as e:
        print(f"[payment-approve] withdrawal #{wid}: {e}")

async def withdrawal_reject_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    wid=int(q.data.split(":")[1])
    w=c.execute("SELECT * FROM withdrawals WHERE id=? AND status='pending'",(wid,)).fetchone()
    if not w:return await q.answer("Already processed.",show_alert=True)
    c.execute("UPDATE withdrawals SET status='rejected' WHERE id=?",(wid,))
    c.execute("UPDATE users SET balance=balance+? WHERE user_id=?",(float(w["total_debited"] or w["amount"]),w["user_id"]))
    db.commit()
    await q.answer("Rejected and refunded.")
    await q.message.edit_reply_markup(reply_markup=None)
    try: await q.message.bot.send_message(w["user_id"],f"Withdrawal #{wid} rejected.\\nRefunded: {w['amount']:.2f} টাকা",reply_markup=um(w["user_id"]))
    except: pass

async def review_file_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    sid=int(q.data.split(":")[1])
    s=c.execute("SELECT * FROM submissions WHERE id=? AND status='pending'",(sid,)).fetchone()
    if not s:return await q.answer("File not found or already processed.",show_alert=True)
    c.execute("INSERT INTO file_reviews(submission_id,status,updated_at,admin_file_message_id,pending_list_message_id) VALUES(?,?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET status='reviewing',updated_at=excluded.updated_at,admin_file_message_id=excluded.admin_file_message_id,pending_list_message_id=0",(sid,"reviewing",now(),q.message.message_id,0))
    db.commit(); state[ADMIN_ID]={"action":"review_accepted","sid":sid}
    await q.answer("Review started")
    prompt=await ans(q.message, f"Submission #{sid}\nReview করার পর Accepted Account সংখ্যা পাঠান।\nTotal: {s['rows_count']}")
    c.execute("UPDATE file_reviews SET review_prompt_message_id=? WHERE submission_id=?",(prompt.message_id,sid));db.commit()


async def reject_file_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    sid=int(q.data.split(":")[1])
    s=c.execute("SELECT * FROM submissions WHERE id=? AND status='pending'",(sid,)).fetchone()
    if not s:return await q.answer("File not found or already processed.",show_alert=True)
    state[ADMIN_ID]={"action":"reject_reason","sid":sid}
    await q.answer()
    await ans(q.message, f"Submission #{sid}\nReject করার কারণ লিখুন।")

async def cleanup_review_submission(bot, sid, review_row=None, submission_row=None):
    """Delete the reviewed upload from storage, Telegram receiver message, and DB.
    User account data is intentionally untouched.
    """
    if submission_row is None:
        submission_row = c.execute("SELECT * FROM submissions WHERE id=?", (sid,)).fetchone()
    if review_row is None:
        review_row = c.execute("SELECT * FROM file_reviews WHERE submission_id=?", (sid,)).fetchone()
    # Remove the physical uploaded XLSX.
    if submission_row:
        try:
            path = submission_row["file_path"]
            if path:
                Path(path).unlink(missing_ok=True)
        except Exception as e:
            print(f"review file delete error #{sid}: {e}")
    # Remove the copy posted to the receiver channel, when possible.
    receiver_chat = setting("receiver_channel").strip()
    if submission_row and receiver_chat:
        try:
            mid = int(submission_row["receiver_message_id"] or 0)
        except Exception:
            mid = 0
        if mid:
            try:
                await bot.delete_message(receiver_chat, mid)
            except Exception as e:
                print(f"receiver message delete error #{sid}: {e}")
    # Remove review UI messages from the admin/receiver chat.
    if review_row:
        for chat_id, mid in [
            (str(ADMIN_ID), review_row["admin_file_message_id"]),
            (str(ADMIN_ID), review_row["review_prompt_message_id"]),
            (str(ADMIN_ID), review_row["pending_list_message_id"]),
        ]:
            try:
                if chat_id and mid:
                    await bot.delete_message(chat_id, int(mid))
            except Exception:
                pass
    # The submission/review rows are temporary review data, not user data.
    c.execute("DELETE FROM file_reviews WHERE submission_id=?", (sid,))
    c.execute("DELETE FROM submissions WHERE id=?", (sid,))
    db.commit()


async def review_confirm_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    sid=int(q.data.split(":")[1])
    r=c.execute("SELECT * FROM file_reviews WHERE submission_id=? AND status='reviewing'",(sid,)).fetchone()
    s=c.execute("SELECT * FROM submissions WHERE id=? AND status='pending'",(sid,)).fetchone()
    if not r or not s:return await q.answer("Review not found or already processed.",show_alert=True)
    c.execute("UPDATE submissions SET status='approved' WHERE id=?",(sid,))
    c.execute("UPDATE users SET balance=balance+?,total_income=total_income+?,today_income=today_income+? WHERE user_id=?",
              (r["final_amount"],r["final_amount"],r["final_amount"],s["user_id"]))
    c.execute("UPDATE file_reviews SET status='approved',updated_at=? WHERE submission_id=?",(now(),sid))
    db.commit();state.pop(ADMIN_ID,None)
    # Hide the reviewed channel/admin messages, but NEVER delete the stored
    # XLSX or the submission record. This keeps approved data available to
    # Generate Main Excel while removing the review UI from the receiver chat.
    review_chat_id = q.message.chat.id
    for mid in (r["admin_file_message_id"], r["pending_list_message_id"], r["review_prompt_message_id"], q.message.message_id):
        try:
            if mid:
                await q.bot.delete_message(review_chat_id,int(mid))
        except Exception:
            pass
    # Keep stored XLSX for Generate Main Excel; only review messages are removed.
    await q.answer("Approved")
    await q.bot.send_message(ADMIN_ID,f"Submission #{sid} approved.\nAccepted: {r['accepted_count']}\nRejected: {r['rejected_count']}\nAdded Balance: {r['final_amount']:.2f}",reply_markup=am())
    try:
        review_complete_text = (
            "{{PE:5203993413346680064}} Report Review Complete\n\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "{{PE:5431420628978641216}} Review Summery\n\n"
            "{{PE:5438605546624132206}} Submitted       : "
            f"{s['rows_count']}\n"
            "{{PE:5951982867255924070}} Accepted        : "
            f"{r['accepted_count']}\n"
            "{{PE:5208928614237623166}} Rejected        : "
            f"{r['rejected_count']}\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "{{PE:5224257782013769471}} Added to Balance : "
            f"{r['final_amount']:.2f}\n"
            "━━━━━━━━━━━━━━━━━━\n\n"
            "{{PE:6235311100879968812}} Your File Has been Reviewed Successfully."
        )
        await sendmsg(
            q.message.bot,
            s["user_id"],
            review_complete_text,
            reply_markup=um(s["user_id"])
        )
    except Exception:
        pass
    await cleanup_review_submission(q.bot, sid, r, s)

async def review_cancel_callback(q):
    if q.from_user.id!=ADMIN_ID:return await q.answer("Not allowed.",show_alert=True)
    state.pop(ADMIN_ID,None)
    await q.answer("Cancelled")
    try: await q.message.delete()
    except: pass

async def approve(m):
    if m.from_user.id!=ADMIN_ID:return
    sid=int(m.text.split()[1]);s=c.execute("SELECT * FROM submissions WHERE id=?",(sid,)).fetchone()
    if not s or s["status"]!="pending":return await ans(m, "Submission not found or already processed.")
    c.execute("UPDATE submissions SET status='approved' WHERE id=?",(sid,));c.execute("UPDATE users SET balance=balance+?,total_income=total_income+?,today_income=today_income+? WHERE user_id=?",(s["estimated_profit"],s["estimated_profit"],s["estimated_profit"],s["user_id"]));db.commit();await ans(m, f"Approved #{sid}",reply_markup=am());await sendmsg(m.bot, s["user_id"],f"Submission #{sid} approved.\nAdded: {s['estimated_profit']:.2f}",reply_markup=um(s["user_id"]));await cleanup_review_submission(m.bot, sid)
async def reject(m):
    if m.from_user.id!=ADMIN_ID:return
    sid=int(m.text.split()[1]);s=c.execute("SELECT * FROM submissions WHERE id=?",(sid,)).fetchone()
    if not s or s["status"]!="pending":return await ans(m, "Submission not found or already processed.")
    c.execute("UPDATE submissions SET status='rejected' WHERE id=?",(sid,));db.commit();await ans(m, f"Rejected #{sid}",reply_markup=am());await sendmsg(m.bot, s["user_id"],f"Submission #{sid} rejected.",reply_markup=um(s["user_id"]));await cleanup_review_submission(m.bot, sid)
async def aw(m):
    if m.from_user.id!=ADMIN_ID:return
    wid=int(m.text.split()[1]);w=c.execute("SELECT * FROM withdrawals WHERE id=?",(wid,)).fetchone()
    if not w or w["status"]!="pending":return await ans(m, "Withdrawal not found or already processed.")
    c.execute("UPDATE withdrawals SET status='approved' WHERE id=?",(wid,));db.commit();await ans(m, f"Withdrawal approved #{wid}",reply_markup=am())
    method = "Bkash" if "bkash" in str(w["category_name"] or "").lower() else "Nogod"
    success_text = (
        "{{PE:6190336264940559752}} WITHDRAWAL SUCCESSFUL\n"
        "━━━━━━━━━━━━━━━━━━━\n"
        f"{{{{PE:6066456029300788292}}}} Amount: {float(w['amount']):.2f} BDT\n"
        f"{{{{PE:6237975191784266396}}}} Method: {method}\n"
        f"{{{{PE:5291913773307667021}}}} Request ID: #{wid}\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        "{{PE:5951982867255924070}} Withdrawal Completed\n"
        "Your payment has been successfully sent.\n"
        "Thank you for using our service. {{PE:5474525960143385880}}\n"
        "━━━━━━━━━━━━━━━━━━━"
    )
    await sendmsg(m.bot, w["user_id"], success_text, reply_markup=um(w["user_id"]))
    await send_payment_approval_report(m.bot,w)
async def rw(m):
    if m.from_user.id!=ADMIN_ID:return
    wid=int(m.text.split()[1]);w=c.execute("SELECT * FROM withdrawals WHERE id=?",(wid,)).fetchone()
    if not w or w["status"]!="pending":return await ans(m, "Withdrawal not found or already processed.")
    c.execute("UPDATE withdrawals SET status='rejected' WHERE id=?",(wid,));c.execute("UPDATE users SET balance=balance+? WHERE user_id=?",(float(w["total_debited"] or w["amount"]),w["user_id"]));db.commit();await ans(m, f"Withdrawal rejected #{wid}; refunded {w['amount']:.2f}",reply_markup=am());await sendmsg(m.bot, w["user_id"],f"Withdrawal #{wid} rejected.\nRefunded: {float(w['total_debited'] or w['amount']):.2f} টাকা (Withdrawal + Fee)",reply_markup=um(w["user_id"]))
async def merge(m):
    if m.from_user.id!=ADMIN_ID:return
    rows=c.execute("SELECT * FROM submissions WHERE status!='rejected' ORDER BY id").fetchall()
    if not rows:return await ans(m, "কোনো uploaded file নেই.",reply_markup=am())
    wb=Workbook();ws=wb.active;ws.title="Main Data";header_written=False
    header_words={"username","user name","userid","user id","email","e-mail","password","pass","cookie","cookies","account","accounts","mail","login","phone","number","mobile","id"}
    for s in rows:
        p=Path(s["file_path"])
        if not p.exists():continue
        try:
            x=load_workbook(p,read_only=True,data_only=True);sh=x.active
            data_rows=[]
            for row in sh.iter_rows(values_only=True):
                v=list(row)
                if any(a is not None and str(a).strip() for a in v):
                    data_rows.append(v)
            x.close()
            if not data_rows: continue
            first=data_rows[0]
            first_norm=[str(v).strip().lower() if v is not None else "" for v in first]
            header_hits=sum(1 for v in first_norm if v in header_words)
            looks_like_header=header_hits>=2 or (header_hits==1 and len(first_norm)>1 and first_norm[0] in {"username","user name","email","e-mail","account","accounts","userid","user id"})
            if looks_like_header:
                if not header_written:
                    ws.append(first);header_written=True
                data_rows=data_rows[1:]
            elif not header_written:
                # No header was supplied; preserve every row exactly as provided.
                header_written=True
            for v in data_rows:
                ws.append(v)
        except Exception:
            pass
    b=io.BytesIO();wb.save(b);b.seek(0)
    await senddoc(m.bot, ADMIN_ID,BufferedInputFile(b.getvalue(),filename=f"main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),caption="Main Excel তৈরি হয়েছে.")


async def pending_files_direct(m):
    """Dedicated Pending Files route; never depends on the broad admin handler."""
    if m.from_user.id != ADMIN_ID:
        return
    try:
        rows = c.execute(
            "SELECT * FROM submissions WHERE LOWER(TRIM(COALESCE(status,'')))='pending' ORDER BY id DESC"
        ).fetchall()
    except Exception as e:
        print(f"pending files query error: {e}")
        return await ans(m, "Pending Files খুলতে সমস্যা হয়েছে। আবার চেষ্টা করুন।", reply_markup=am())

    if not rows:
        return await ans(m, "📂 Pending Files\n\nকোনো pending file নেই।", reply_markup=am())

    sent = 0
    unavailable = 0
    for s in rows:
        try:
            caption = (
                f"New Account File\n\nSubmission ID: {s['id']}\n"
                f"Username: @{s['username'] or 'N/A'}\nUser ID: {s['user_id']}\n"
                f"Work: {s['category_name'] or 'N/A'}\nAccounts: {int(s['rows_count'] or 0)}\n"
                f"Price: {float(s['price'] or 0):.2f}\n"
                f"Estimated Profit: {float(s['estimated_profit'] or 0):.2f}\n"
                f"File: {s['file_name'] or 'N/A'}\nTime: {s['created_at'] or 'N/A'}"
            )
        except Exception:
            caption = f"New Account File\n\nSubmission ID: {s['id']}"

        # Plain inline buttons: avoid optional Premium fields breaking this route.
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            _InlineKeyboardButton(text="Review File", callback_data=f"review:{s['id']}"),
            _InlineKeyboardButton(text="Reject File", callback_data=f"reject_file:{s['id']}")
        ]])

        delivered = False
        msg = None
        try:
            fid = s['file_id']
        except Exception:
            fid = None
        if fid:
            try:
                msg = await m.bot.send_document(ADMIN_ID, fid, caption=caption, reply_markup=kb)
                delivered = True
            except Exception as e:
                print(f"pending file #{s['id']} file_id error: {e}")

        if not delivered:
            try:
                fp = Path(s['file_path'])
                if fp.exists():
                    msg = await senddoc(
                        m.bot, ADMIN_ID,
                        BufferedInputFile(fp.read_bytes(), filename=s['file_name'] or 'file.xlsx'),
                        caption=caption, reply_markup=kb
                    )
                    delivered = True
            except Exception as e:
                print(f"pending file #{s['id']} local file error: {e}")

        if not delivered:
            try:
                rid = int(s['receiver_message_id'] or 0)
            except Exception:
                rid = 0
            receiver = setting('receiver_channel').strip()
            if rid and receiver:
                try:
                    msg = await m.bot.copy_message(
                        chat_id=ADMIN_ID, from_chat_id=receiver, message_id=rid,
                        caption=caption, reply_markup=kb
                    )
                    delivered = True
                except Exception as e:
                    print(f"pending file #{s['id']} receiver copy error: {e}")

        if delivered:
            try:
                c.execute(
                    "INSERT INTO file_reviews(submission_id,status,updated_at,pending_list_message_id) "
                    "VALUES(?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET "
                    "pending_list_message_id=excluded.pending_list_message_id",
                    (s['id'], 'pending', now(), msg.message_id)
                )
                db.commit()
            except Exception as e:
                print(f"pending file #{s['id']} tracking error: {e}")
            sent += 1
        else:
            unavailable += 1
            # Still show the record so the button never appears to do nothing.
            try:
                msg = await ans(m, caption + "\n\n⚠️ File could not be recovered.", reply_markup=kb)
                c.execute(
                    "INSERT INTO file_reviews(submission_id,status,updated_at,pending_list_message_id) "
                    "VALUES(?,?,?,?) ON CONFLICT(submission_id) DO UPDATE SET "
                    "pending_list_message_id=excluded.pending_list_message_id",
                    (s['id'], 'pending', now(), msg.message_id)
                )
                db.commit()
            except Exception as e:
                print(f"pending file #{s['id']} metadata error: {e}")

    suffix = f"\nUnavailable: {unavailable}" if unavailable else ""
    return await ans(m, f"📂 Pending Files\n\nShowing {sent} pending file(s).{suffix}", reply_markup=am())

async def pending_withdrawals_requests_direct(m):
    """Dedicated route for both Pending Withdrawals button labels."""
    if m.from_user.id != ADMIN_ID:
        return
    try:
        rows = c.execute(
            "SELECT id,user_id,username,full_name,category_name,amount,wallet,status,created_at "
            "FROM withdrawals WHERE LOWER(TRIM(COALESCE(status,'')))='pending' ORDER BY id DESC"
        ).fetchall()
    except Exception as e:
        print(f"pending withdrawals query error: {e}")
        return await ans(m, "Pending Withdrawals খুলতে সমস্যা হয়েছে। আবার চেষ্টা করুন।", reply_markup=am())

    if not rows:
        return await ans(m, "Pending Withdrawals: 0\n\nকোনো pending withdrawal নেই।", reply_markup=am())

    sent = 0
    for x in rows:
        try:
            amount = float(x['amount'] or 0)
        except Exception:
            amount = 0.0
        text = (
            f"Withdrawal #{x['id']}\n\n"
            f"Account Name: {x['full_name'] or 'N/A'}\n"
            f"Username: @{x['username'] or 'N/A'}\n"
            f"User ID: {x['user_id']}\n"
            f"Category: {x['category_name'] or 'N/A'}\n"
            f"Amount: {amount:.2f}\n"
            f"Payment Number: {x['wallet'] or 'N/A'}\n"
            f"Time: {x['created_at'] or 'N/A'} UTC"
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            _InlineKeyboardButton(text="Approve", callback_data=f"wapprove:{x['id']}"),
            _InlineKeyboardButton(text="Reject + Refund", callback_data=f"wreject:{x['id']}")
        ]])
        try:
            await ans(m, text, reply_markup=kb)
            sent += 1
        except Exception as e:
            print(f"pending withdrawal #{x['id']} send error: {e}")
    return await ans(m, f"Pending Withdrawals: {sent}", reply_markup=am())

async def priority_admin_navigation(m):
    if m.from_user.id != ADMIN_ID:
        return
    t = clean_button_text(m.text)
    if t == "Pending Withdrawals":
        return await pending_withdrawals_menu(m)
    if t == "Pending Files":
        return await admin(m)
    if t == "Upload Report File":
        return await report_admin_action(m)
    if t == "Support":
        return await support(m)

dp=Dispatcher()
dp.message.outer_middleware(MaintenanceMiddleware())
dp.callback_query.outer_middleware(MaintenanceMiddleware())
# Register critical Admin Panel navigation before broad admin/menu handlers.
# V13: Working Pending Withdrawals handler based only on the verified reference flow.
async def pending_withdrawals_direct_v13(m):
    """Verified admin route for both pending-withdrawal button labels."""
    if m.from_user.id != ADMIN_ID:
        return

    try:
        rows = c.execute(
            "SELECT id,user_id,username,full_name,category_name,amount,wallet,status,created_at "
            "FROM withdrawals "
            "WHERE LOWER(TRIM(COALESCE(status,'')))='pending' "
            "ORDER BY id DESC"
        ).fetchall()
    except Exception as e:
        print(f"[pending-withdrawals] database error: {e}")
        return await m.answer(
            "Pending Withdrawals\n\nDatabase error while loading requests.",
            reply_markup=pending_admin_kb()
        )

    if not rows:
        return await m.answer(
            "Pending Withdrawals: 0\n\nNo pending withdrawal requests.",
            reply_markup=pending_admin_kb()
        )

    sent = 0
    for x in rows:
        try:
            amount = float(x["amount"] or 0)
        except Exception:
            amount = 0.0

        wid = x["id"]
        msg = (
            f"Withdrawal #{wid}\n\n"
            f"Account Name: {x['full_name'] or 'N/A'}\n"
            f"Username: @{x['username'] or 'N/A'}\n"
            f"User ID: {x['user_id']}\n"
            f"Category: {x['category_name'] or 'N/A'}\n"
            f"Amount: {amount:.2f}\n"
            f"Payment Number: {x['wallet'] or 'N/A'}\n"
            f"Time: {x['created_at'] or 'N/A'}"
        )
        markup = InlineKeyboardMarkup(inline_keyboard=[[
            _InlineKeyboardButton(text="Approve", callback_data=f"wapprove:{wid}"),
            _InlineKeyboardButton(text="Reject", callback_data=f"wreject:{wid}")
        ]])

        try:
            await m.answer(msg, reply_markup=markup)
            sent += 1
        except Exception as e:
            print(f"[pending-withdrawals] send error for #{wid}: {e}")

    return await m.answer(
        f"Pending Withdrawals: {sent}",
        reply_markup=pending_admin_kb()
    )

# Register this BEFORE every broad admin text handler.
dp.message.register(
    pending_withdrawals_direct_v13,
    F.from_user.id == ADMIN_ID,
    lambda m: clean_button_text(m.text) in {"Pending Withdrawals", "Pending Withdrawal Requests"}
)
dp.message.register(start,CommandStart())
dp.message.register(userpanel,Command("userpanel"))
dp.message.register(database_upload_doc, F.document, F.from_user.id == ADMIN_ID, lambda m: state.get(ADMIN_ID, {}).get("action") == "database_upload")
dp.message.register(report_doc, F.document, F.from_user.id == ADMIN_ID, lambda m: state.get(ADMIN_ID, {}).get("action") == "upload_report")
dp.message.register(doc, F.document)
dp.callback_query.register(review_file_callback,F.data.startswith("review:"))
dp.callback_query.register(reject_file_callback,F.data.startswith("reject_file:"))
dp.callback_query.register(review_confirm_callback,F.data.startswith("review_confirm:"))
dp.callback_query.register(review_cancel_callback,F.data.startswith("review_cancel:"))
dp.callback_query.register(withdrawal_add_method_callback,F.data.startswith("waddmethod:"))
dp.callback_query.register(withdrawal_edit_callback,F.data.startswith("wedit:"))
dp.callback_query.register(withdrawal_delete_callback,F.data.startswith("wdelete:"))
dp.callback_query.register(withdrawal_toggle_callback,F.data.startswith("wtoggle:"))
dp.callback_query.register(withdrawal_approve_callback,F.data.startswith("wapprove:"))
dp.callback_query.register(withdrawal_reject_callback,F.data.startswith("wreject:"))
dp.callback_query.register(report_category_callback,F.data.startswith("report:"))
dp.callback_query.register(force_join_callback,F.data=="check_force_join")
dp.message.register(force_join_management,F.text=="Force Join")
# Critical admin navigation must be registered before broad admin/config handlers.
dp.message.register(
    maintenance_admin_action,
    F.from_user.id == ADMIN_ID,
    lambda m: clean_button_text(m.text) in {"Maintenance", "Maintenance ON", "Maintenance OFF", "Back to Admin Panel"}
)
dp.message.register(
    database_admin_action,
    F.from_user.id == ADMIN_ID,
    lambda m: clean_button_text(m.text) in {"Database", "Download Database", "Upload Database", "ডিলিট অল ডাটা", "YES", "CANCEL"} or state.get(ADMIN_ID, {}).get("action") in {"database_upload", "database_delete_confirm"}
)
dp.message.register(
    pending_files_direct,
    F.from_user.id == ADMIN_ID,
    lambda m: clean_button_text(m.text) == "Pending Files"
)
dp.message.register(
    withdrawal_admin_action,
    F.from_user.id==ADMIN_ID,
    lambda m: (
        m.text in {"Add Withdrawal Category","Delete Category","Withdrawal Category List","Pending Withdrawal Requests","Pending Withdrawals","Back to Admin Panel"}
        or state.get(ADMIN_ID, {}).get("action", "").startswith("w_")
    )
)
dp.message.register(support_admin_action,F.from_user.id==ADMIN_ID, lambda m: (
    m.text in {"Add Support Button","Delete Support Button","Support List","Back to Admin Panel"}
    or state.get(ADMIN_ID, {}).get("action", "").startswith("support_btn_")
))
dp.message.register(work_menu_action,F.from_user.id==ADMIN_ID, lambda m: (m.text in {"Add Work","Edit Work","Disable Work","Enable Work","Disable Notice","Work List","Back to Admin Panel"} or state.get(ADMIN_ID, {}).get("action") in {"disable_notice","work_name","work_price","work_edit_id","work_edit_name","work_edit_price","work_toggle_id"}))
dp.message.register(force_join_admin_action,F.from_user.id==ADMIN_ID, F.text.in_({"Add Force Join Channel","Force Join Channel List","Remove Force Join Channel","Back to Admin Panel"}))
dp.message.register(user_management,F.text=="User Management")
dp.message.register(user_mgmt_action,F.from_user.id==ADMIN_ID, F.text.in_({"All User List","Block User","Blocked User","Add Balance","Remove Balance","Back to Admin Panel"}))
# Built-in User Panel actions take priority over dynamically-created custom buttons.
dp.message.register(account,F.text, lambda m: clean_button_text(m.text)=="Account")
dp.message.register(report_section,F.text, lambda m: clean_button_text(m.text) in {"Report","Pending Account"})
dp.message.register(sell,F.text, lambda m: clean_button_text(m.text)=="Sell Account")
dp.message.register(withdraw,F.text, lambda m: clean_button_text(m.text)=="Withdrawal")
dp.message.register(prices,F.text, lambda m: clean_button_text(m.text)=="Price List")
dp.message.register(support,F.text, lambda m: clean_button_text(m.text)=="Support")

# Dynamic buttons are checked after built-in actions, so an added button named
# "Support" can never break the main Support button.
dp.message.register(handle_custom_button, F.text, lambda m: bool(c.execute("SELECT 1 FROM custom_buttons WHERE name=? AND active=1",(clean_button_text(m.text),)).fetchone()))

# Admin config handler must only claim messages while one of its own setup
# flows is active. A broad admin text handler would swallow review answers
# (e.g. the accepted-account count sent from the receiver channel) and the
# Admin Panel navigation itself.
dp.message.register(
    new_button_action,
    F.from_user.id==ADMIN_ID,
    lambda m: (
        m.text in {"New Button","Add New Button","View Added Button","Edit Added Button","Delete Button","Back to Admin Panel"}
        or state.get(ADMIN_ID, {}).get("action") in {"edit_button_id","edit_button_name","edit_button_target","delete_button_id"}
    )
)
dp.message.register(
    admin_config_action,
    F.from_user.id==ADMIN_ID,
    lambda m: state.get(ADMIN_ID, {}).get("action") in {"add_button_name", "add_button_target", "receiver_channel"}
)
dp.message.register(report_admin_action,F.from_user.id==ADMIN_ID, lambda m: clean_button_text(m.text) in {"Upload Report File","Future Reports"} or state.get(ADMIN_ID,{}).get("action") in {"upload_report_category","upload_report"})
async def pending_withdrawals_menu(m):
    if m.from_user.id != ADMIN_ID:
        return
    try:
        rows = c.execute(
            "SELECT id,user_id,username,full_name,category_name,amount,wallet,status,created_at "
            "FROM withdrawals WHERE LOWER(TRIM(COALESCE(status,'')))='pending' ORDER BY id DESC"
        ).fetchall()
    except Exception as e:
        print(f"pending withdrawals query error: {e}")
        return await ans(m, "Pending Withdrawals খুলতে সমস্যা হয়েছে। আবার চেষ্টা করুন।", reply_markup=am())

    if not rows:
        return await ans(m, "Pending Withdrawals: 0\n\nকোনো pending withdrawal নেই।", reply_markup=am())

    sent = 0
    for x in rows:
        try:
            amount = float(x["amount"] or 0)
        except Exception:
            amount = 0.0
        text = (
            f"Withdrawal #{x['id']}\n\n"
            f"Account Name: {x['full_name'] or 'N/A'}\n"
            f"Username: @{x['username'] or 'N/A'}\n"
            f"User ID: {x['user_id']}\n"
            f"Category: {x['category_name'] or 'N/A'}\n"
            f"Amount: {amount:.2f}\n"
            f"Payment Number: {x['wallet'] or 'N/A'}\n"
            f"Time: {x['created_at'] or 'N/A'} UTC"
        )
        # Use raw inline buttons here so a bad custom-emoji/style field can
        # never prevent the Pending Withdrawals screen from opening.
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Approve", callback_data=f"wapprove:{x['id']}"),
            InlineKeyboardButton(text="Reject", callback_data=f"wreject:{x['id']}")
        ]])
        try:
            await ans(m, text, reply_markup=kb)
            sent += 1
        except Exception as e:
            print(f"pending withdrawal #{x['id']} send error: {e}")

    return await ans(m, f"Pending Withdrawals: {sent}", reply_markup=am())


async def admin_navigation(m):
    if m.from_user.id != ADMIN_ID:
        return
    t = clean_button_text(m.text)
    # Route navigation explicitly so no broad admin handler can swallow it.
    if t == "Pending Withdrawals":
        return await pending_withdrawals_menu(m)
    if t == "Pending Files":
        return await admin(m)
    if t == "Support":
        return await support(m)
    if t in {"Upload Report File", "Future Reports"}:
        return await report_admin_action(m)
    return await admin(m)

dp.message.register(admin,F.from_user.id==ADMIN_ID)
# Generic user text handler must run before the broad category handlers so
# amount/wallet input is not swallowed by another text handler.
dp.message.register(user_state,F.text)
dp.message.register(sell_category_selection,F.text, F.from_user.id != ADMIN_ID)
dp.message.register(withdrawal_category_selection,F.text, F.from_user.id != ADMIN_ID)


async def report_cleanup_loop():
    while True:
        try:
            cleanup_expired_report()
        except Exception:
            pass
        await asyncio.sleep(300)

async def main():
    b=Bot(TOKEN,default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    asyncio.create_task(report_cleanup_loop())
    await dp.start_polling(b)
if __name__=="__main__":asyncio.run(main())
